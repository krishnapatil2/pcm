import math
import os
import pandas as pd
import zipfile
import cons_header
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter

def safe_float(val):
    """Convert to float safely, treating blanks/NA as 0.0"""
    if pd.isna(val):
        return 0
    try:
        num = float(val)
        # Round .5 and above up, below .5 down
        return math.floor(num + 0.5)
    except Exception:
        return 0

def safe_float_no_round(val):
    """Convert to float safely without rounding, treating blanks/NA as 0.0"""
    if pd.isna(val):
        return 0.0
    try:
        return float(val)
    except Exception:
        return 0.0
    
def read_file(filepath: str, header: int = 0, custom_header: list = None) -> pd.DataFrame:
    """
    Reads a file (CSV, XLS, XLSX) into a pandas DataFrame.
    
    Args:
        filepath (str): Path to the input file
        header (int): Row number to use as column names (default=0)
        custom_header (list): If provided, select only these columns after loading
    
    Returns:
        pd.DataFrame: Loaded data
    """
    # Load with original headers from file
    if filepath.lower().endswith(".csv"):
        df = pd.read_csv(filepath, header=header)
    elif filepath.lower().endswith(".xlsx"):
        df = pd.read_excel(filepath, engine="openpyxl", header=header)
    elif filepath.lower().endswith(".xls"):
        df = pd.read_excel(filepath, engine="xlrd", header=header)
    else:
        raise ValueError("Unsupported file format. Supported: .csv, .xls, .xlsx")
    
    # Strip spaces from column headers
    df.columns = df.columns.str.strip()

    # If custom_header is provided, select only those columns
    if custom_header:
        custom_header = [col.strip() for col in custom_header]
        missing = [col for col in custom_header if col not in df.columns]
        if missing:
            raise ValueError(f"These columns are missing in file: {missing}")
        df = df[custom_header]
    
    return df


def build_dict(file_path: str, key_cols: list, value_cols: dict, filter_col: str = None, filter_val=None, no_round_cols: list = None) -> dict:
    """
    Reads CSV, applies optional filter, builds dictionary for manual updates.

    Args:
        file_path (str): Path to CSV file
        key_cols (list): Column names to form dictionary key (tuple if >1)
        value_cols (dict): { "Output Column Name": "Source Column Name" }
        filter_col (str, optional): Column name for filter
        filter_val (any, optional): Value to filter rows
        no_round_cols (list, optional): List of output column names that should not be rounded

    Returns:
        dict: { (keys): { "Output Column": value, ... } }
    """
    df = pd.read_csv(file_path, header=0)
    df.columns = df.columns.str.strip()

    # Apply filter if provided
    if filter_col and filter_val is not None:
        df = df[df[filter_col] == filter_val]

    # Default columns that should not be rounded
    if no_round_cols is None:
        no_round_cols = []
    
    lookup_dict = {}
    for _, row in df.iterrows():
        # key = tuple(str(row[col]).strip() for col in key_cols)
        key = tuple(
                        str(int(row[col])) if col == "BrkrOrCtdnPtcptId" and isinstance(row[col], float) and row[col].is_integer()
                        else str(row[col]).strip()
                        for col in key_cols
                    )        

        value_dict = {}
        for out_col, src_col in value_cols.items():
            val = row[src_col]

            # Replace NaN/None with 0 or 0.0
            if pd.isna(val):
                val = ''

            # Use no-round function for specified columns, otherwise use safe_float
            if out_col in no_round_cols:
                value_dict[out_col] = safe_float_no_round(val)
            else:
                value_dict[out_col] = safe_float(val)

        lookup_dict[key] = value_dict

    return lookup_dict


def _format_summary_sheet(worksheet, summary_df, thin_border):
    """Helper function to format a summary sheet with borders and styling."""
    header_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Format header row
    for col in range(1, len(summary_df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_alignment
    
    # Format data rows
    for row in range(2, len(summary_df) + 2):
        for col in range(1, len(summary_df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
    
    # Apply comma style formatting to summary numeric columns
    summary_comma_columns = [
        "Physical Settlement Obligation",
        "Physical Settlement STT",
        "Physical Settlement SD",
        "Total Obligation"
    ]
    
    # Find column indices for comma style columns
    for col_idx, col_name in enumerate(summary_df.columns, 1):
        if col_name in summary_comma_columns:
            # Apply comma format to all data rows (skip header row 1)
            for row_idx in range(2, len(summary_df) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
    
    # Auto-adjust column widths
    for col_idx, col_name in enumerate(summary_df.columns, 1):
        column_letter = worksheet.cell(row=1, column=col_idx).column_letter
        max_length = len(str(col_name))
        for row in range(2, len(summary_df) + 2):
            cell_value = worksheet.cell(row=row, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        worksheet.column_dimensions[column_letter].width = max(max_length + 2, 15)


def create_segregated_file_with_summary(obligation_path: str,
                                        stt_path: str,
                                        stamp_duty_path: str,
                                        output_path: str,
                                        column_name: str = "BrkrOrCtdnPtcptId",
                                        custom_header: list = None,
                                        update_dicts: list = None):
    """
    Creates separate Excel files for each BrkrOrCtdnPtcptId and a Summary file, then zips them all.
    
    Args:
        obligation_path (str): Path to obligation file
        stt_path (str): Path to STT CSV file
        stamp_duty_path (str): Path to Stamp Duty CSV file
        output_path (str): Path to save ZIP file containing all Excel files
        column_name (str): Column to group by (default: BrkrOrCtdnPtcptId)
        custom_header (list, optional): Custom header columns to read
        update_dicts (list, optional): List of dictionaries to update values from
    """
    # Read obligation file
    df = read_file(obligation_path, custom_header=custom_header)
    
    if column_name not in df.columns:
        raise ValueError(f"Column '{column_name}' not found in file. Found: {df.columns.tolist()}")
    
    # Add extra columns at the end in the specified order
    for col in cons_header.EXTRA_COLUMNS:
        if col not in df.columns:
            df[col] = pd.NA
    
    # Replace blanks/NaN in grouping column with "Blank"
    df[column_name] = df[column_name].fillna("").astype(str).str.strip()
    df[column_name] = df[column_name].replace({"": "Blank"})
    
    # Create output directory if it doesn't exist
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    
    # Create temporary directory for individual files
    temp_dir = os.path.join(output_dir, "temp_segregated_files")
    os.makedirs(temp_dir, exist_ok=True)
    
    try:
        # Prepare summary data
        summary_data = []
        all_participants = set()
        
        # Get all unique participants from obligation file
        all_participants.update(df[column_name].unique())
        
        # Calculate Physical Obligation for each participant
        obligation_summary = {}
        for participant, group in df.groupby(column_name):
            cmltv_buy_sum = safe_float_no_round(group["CmltvBuyAmt"].sum()) if "CmltvBuyAmt" in group.columns else 0.0
            cmltv_sell_sum = safe_float_no_round(group["CmltvSellAmt"].sum()) if "CmltvSellAmt" in group.columns else 0.0
            physical_obligation = cmltv_sell_sum - cmltv_buy_sum
            obligation_summary[participant] = physical_obligation
        
        # Read and filter STT data (RptHdr == 20)
        stt_data = pd.read_csv(stt_path)
        stt_filtered = stt_data[stt_data["RptHdr"] == 20].copy()
        # Convert BrkrOrCtdnPtcptId to string for consistent matching
        stt_filtered["BrkrOrCtdnPtcptId"] = stt_filtered["BrkrOrCtdnPtcptId"].astype(str).str.strip()
        stt_summary = stt_filtered.groupby("BrkrOrCtdnPtcptId")["TtlTaxs"].sum().to_dict()
        all_participants.update(stt_summary.keys())
        
        # Read and filter Stamp Duty data (RptHdr == 20)
        stamp_duty_data = pd.read_csv(stamp_duty_path)
        stamp_duty_filtered = stamp_duty_data[stamp_duty_data["RptHdr"] == 20].copy()
        # Convert BrkrOrCtdnPtcptId to string for consistent matching
        stamp_duty_filtered["BrkrOrCtdnPtcptId"] = stamp_duty_filtered["BrkrOrCtdnPtcptId"].astype(str).str.strip()
        sd_summary = stamp_duty_filtered.groupby("BrkrOrCtdnPtcptId")["StmpDtyAmt"].sum().to_dict()
        all_participants.update(sd_summary.keys())
        
        # Build summary data - convert all participants to strings for consistency
        # Also convert keys in dictionaries to strings
        obligation_summary_str = {str(k).strip(): v for k, v in obligation_summary.items()}
        stt_summary_str = {str(k).strip(): v for k, v in stt_summary.items()}
        sd_summary_str = {str(k).strip(): v for k, v in sd_summary.items()}
        
        # Build summary data and create lookup dictionary
        summary_lookup = {}
        for participant in sorted(all_participants):
            participant_str = str(participant).strip()
            summary_row = {
                "Client Code": participant_str,
                "Physical Settlement Obligation": obligation_summary_str.get(participant_str, 0.0),
                "Physical Settlement STT": safe_float_no_round(stt_summary_str.get(participant_str, 0.0)),
                "Physical Settlement SD": safe_float_no_round(sd_summary_str.get(participant_str, 0.0))
            }
            summary_data.append(summary_row)
            summary_lookup[participant_str] = summary_row
        
        # Create individual Excel files for each participant
        participant_files = []
        for key, subset in df.groupby(column_name):
            subset = subset.copy()
            
            # Update values from each update_dict if key matches
            for i, row in subset.iterrows():
                val = row["FinInstrmId"]
                if pd.isna(val):
                    val_float = None
                elif str(val).strip() == "":
                    val_float = None
                else:
                    val_float = float(val)
                    val_str = str(val_float)
                
                match_key = (row["BrkrOrCtdnPtcptId"], row["TckrSymb"], val_str)
                
                if update_dicts:
                    for update_dict in update_dicts:
                        if match_key in update_dict:
                            for col, val in update_dict[match_key].items():
                                subset.at[i, col] = val
            
            # Do calculations for each row
            for i, row in subset.iterrows():
                buystamp = row.get("Buy Stamp Duty")
                sellstamp = row.get("Sell Stamp Duty")
                
                if pd.isna(buystamp) or str(buystamp).strip() == "":
                    subset.at[i, "Buy Stamp Duty"] = 0
                if pd.isna(sellstamp) or str(sellstamp).strip() == "":
                    subset.at[i, "Sell Stamp Duty"] = 0
                
                buy_stamp = subset.at[i, "Buy Stamp Duty"]
                sell_stamp = subset.at[i, "Sell Stamp Duty"]
                
                cmltv_buy_val = row.get("CmltvBuyAmt")
                if pd.isna(cmltv_buy_val) or str(cmltv_buy_val).strip() == "":
                    cmltv_buy = 0.0
                else:
                    cmltv_buy = safe_float_no_round(cmltv_buy_val)
                
                cmltv_sell_val = row.get("CmltvSellAmt")
                if pd.isna(cmltv_sell_val) or str(cmltv_sell_val).strip() == "":
                    cmltv_sell = 0.0
                else:
                    cmltv_sell = safe_float_no_round(cmltv_sell_val)
                
                buy_stt = row.get("Buy STT")
                sell_stt = row.get("Sell STT")
                
                buy_payable = cmltv_buy + buy_stt + buy_stamp
                sell_receivable = cmltv_sell - sell_stt - sell_stamp
                net_receivable_payable = sell_receivable - buy_payable
                
                subset.at[i, "Buy Payable Amount"] = buy_payable
                subset.at[i, "Sell Receivable Amount"] = sell_receivable
                subset.at[i, "Net Receivable \\ Payable"] = net_receivable_payable
            
            # Add Total row for "Net Receivable \ Payable"
            if "Net Receivable \\ Payable" in subset.columns:
                total_value = subset["Net Receivable \\ Payable"].apply(lambda x: safe_float_no_round(x) if pd.notna(x) else 0.0).sum()
                total_row = {col: "" for col in subset.columns}
                cols = list(subset.columns)
                net_idx = cols.index("Net Receivable \\ Payable")
                if net_idx > 0:
                    total_row[cols[net_idx - 1]] = "Total"
                total_row["Net Receivable \\ Payable"] = total_value
                subset = pd.concat([subset, pd.DataFrame([total_row])], ignore_index=True)
            
            # Create filename for this participant
            safe_filename = str(key)[:50].replace("/", "_").replace("\\", "_") or "Blank"
            participant_file = os.path.join(temp_dir, f"{safe_filename}.xlsx")
            participant_files.append(participant_file)
            
            # Get this participant's summary row
            participant_str = str(key).strip()
            participant_summary_row = summary_lookup.get(participant_str)
            
            # Write to Excel file
            with pd.ExcelWriter(participant_file, engine="openpyxl") as writer:
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Create Summary sheet first (so it appears as first sheet)
                if participant_summary_row:
                    participant_summary_df = pd.DataFrame([participant_summary_row])
                else:
                    # Create empty summary if not found (shouldn't happen, but just in case)
                    participant_summary_df = pd.DataFrame([{
                        "Client Code": participant_str,
                        "Physical Settlement Obligation": 0.0,
                        "Physical Settlement STT": 0.0,
                        "Physical Settlement SD": 0.0
                    }])
                
                # Calculate Total = Physical Settlement Obligation - Physical Settlement STT - Physical Settlement SD
                physical_obligation = participant_summary_df.iloc[0]["Physical Settlement Obligation"]
                physical_stt = participant_summary_df.iloc[0]["Physical Settlement STT"]
                physical_sd = participant_summary_df.iloc[0]["Physical Settlement SD"]
                total_value = safe_float_no_round(physical_obligation - physical_stt - physical_sd)
                
                # Add Total Obligation column with calculated value
                participant_summary_df["Total Obligation"] = total_value
                
                participant_summary_df.to_excel(writer, sheet_name="Summary", index=False)
                
                # Format Summary sheet
                workbook = writer.book
                ws_summary = workbook["Summary"]
                _format_summary_sheet(ws_summary, participant_summary_df, thin_border)
                
                # Write Details sheet second (so it appears as second sheet)
                subset.to_excel(writer, sheet_name="Details", index=False)
                
                # Add borders and formatting for Details sheet
                worksheet = writer.sheets["Details"]
                
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = thin_border
                
                # Apply comma style formatting to specific columns in Details sheet
                comma_style_columns = [
                    "Buy STT",
                    "Sell STT",
                    "Sell Stamp Duty",
                    "Buy Stamp Duty",
                    "Buy Payable Amount",
                    "Sell Receivable Amount",
                    "Net Receivable \\ Payable"
                ]
                
                # Find column indices for comma style columns
                for col_idx, col_name in enumerate(subset.columns, 1):
                    if col_name in comma_style_columns:
                        # Apply comma format to all data rows (skip header row 1)
                        for row_idx in range(2, len(subset) + 2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            if cell.value is not None and isinstance(cell.value, (int, float)):
                                cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                
                # Auto-adjust column widths for Details sheet
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = max(max_length + 2, 12)
                    if column_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
                        header_cell = worksheet.cell(row=1, column=column[0].column)
                        if header_cell.value in cons_header.EXTRA_COLUMNS:
                            adjusted_width = max(adjusted_width, 18)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create Summary Excel file
        summary_df = pd.DataFrame(summary_data)
        
        # Calculate Total Obligation = Physical Settlement Obligation - Physical Settlement STT - Physical Settlement SD for each row
        summary_df["Total Obligation"] = summary_df.apply(
            lambda row: safe_float_no_round(row["Physical Settlement Obligation"] - row["Physical Settlement STT"] - row["Physical Settlement SD"]),
            axis=1
        )
        
        summary_file = os.path.join(temp_dir, "Summary.xlsx")
        
        with pd.ExcelWriter(summary_file, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            
            # Format Summary sheet
            workbook = writer.book
            ws_summary = workbook["Summary"]
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            _format_summary_sheet(ws_summary, summary_df, thin_border)
        
        # Create ZIP file with all Excel files
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Add Summary file
            zipf.write(summary_file, "Summary.xlsx")
            
            # Add all participant files
            for participant_file in participant_files:
                arcname = os.path.basename(participant_file)
                zipf.write(participant_file, arcname)
        
    finally:
        # Clean up temporary directory
        import shutil
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)


def segregate_excel_by_column(excel_path: str,
                                output_path: str,
                                column_name: str = "BrkrOrCtdnPtcptId", 
                                custom_header: list = None,
                                update_dicts: list = None,
                                stt_path: str = None,
                                stamp_duty_path: str = None):
    """
    Segregates Excel data by unique values of a given column 
    and writes into multiple sheets.
    
    Args:
        excel_path (str): Path to Excel file
        output_path (str): Path to save segregated Excel file
        column_name (str): Column to group by (default: BrkrOrCtdnPtcptId)
        custom_header (list, optional): Custom header columns to read
        update_dicts (list, optional): List of dictionaries to update values from
        stt_path (str, optional): Path to STT CSV file for summary sheet
        stamp_duty_path (str, optional): Path to Stamp Duty CSV file for summary sheet
    """
    df = read_file(excel_path, custom_header=custom_header)  # use your read function
    # print("name:", column_name)
    # print("Unique BrkrOrCtdnPtcptId:", df[column_name].unique())

    if column_name not in df.columns:
        raise ValueError(f"Column '{column_name}' not found in file. Found: {df.columns.tolist()}")

    # Add extra columns at the end in the specified order
    for col in cons_header.EXTRA_COLUMNS:
        if col not in df.columns:
            df[col] = pd.NA #  pd.NA  preferred for empty cells in pandas
    
    # Replace blanks/NaN in grouping column with "Blank"
    df[column_name] = df[column_name].fillna("").astype(str).str.strip()
    df[column_name] = df[column_name].replace({"": "Blank"})

    # Create output directory if it doesn't exist
    output_dir = os.path.dirname(output_path)
    if output_dir:  # Only create directory if path has a directory component
        os.makedirs(output_dir, exist_ok=True)
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for key, subset in df.groupby(column_name):
            subset = subset.copy()                
            # Update values from each update_dict if key matches
            for i, row in subset.iterrows():
                val = row["FinInstrmId"]
                if pd.isna(val):        # NaN in pandas
                    val_float = None    # or leave as None
                elif str(val).strip() == "":  # Blank string
                    val_float = None
                else:
                    val_float = float(val)
                    val_str = str(val_float) # convert float to string

                match_key = (row["BrkrOrCtdnPtcptId"], row["TckrSymb"], val_str)

                # print("Matching Key:", match_key)
                for update_dict in update_dicts:
                    if match_key in update_dict:
                        for col, val in update_dict[match_key].items():
                            subset.at[i, col] = val  # just update directly

            # now i want to access here df and access some columns and do some calculations and update those columns
            # Do calculations for each row in the sheet
            for i, row in subset.iterrows():
                buystamp = row.get("Buy Stamp Duty")
                sellstamp = row.get("Sell Stamp Duty")

                # Handle missing stamp duties
                if pd.isna(buystamp) or str(buystamp).strip() == "":
                    subset.at[i, "Buy Stamp Duty"] = 0

                if pd.isna(sellstamp) or str(sellstamp).strip() == "":
                    subset.at[i, "Sell Stamp Duty"] = 0
                
                buy_stamp   = subset.at[i, "Buy Stamp Duty"]
                sell_stamp  = subset.at[i, "Sell Stamp Duty"]

                # Get CmltvBuyAmt and CmltvSellAmt without rounding
                cmltv_buy_val = row.get("CmltvBuyAmt")
                if pd.isna(cmltv_buy_val) or str(cmltv_buy_val).strip() == "":
                    cmltv_buy = 0.0
                else:
                    cmltv_buy = safe_float_no_round(cmltv_buy_val)
                
                cmltv_sell_val = row.get("CmltvSellAmt")
                if pd.isna(cmltv_sell_val) or str(cmltv_sell_val).strip() == "":
                    cmltv_sell = 0.0
                else:
                    cmltv_sell = safe_float_no_round(cmltv_sell_val)

                buy_stt     = row.get("Buy STT")
                sell_stt    = row.get("Sell STT")

                buy_payable = cmltv_buy + buy_stt + buy_stamp
                sell_receivable = cmltv_sell - sell_stt - sell_stamp
                net_receivable_payable = sell_receivable - buy_payable

                subset.at[i, "Buy Payable Amount"] = buy_payable
                subset.at[i, "Sell Receivable Amount"] = sell_receivable
                subset.at[i, "Net Receivable \\ Payable"] = net_receivable_payable
            
            # ======================
            # Add Total row for "Net Receivable \ Payable"
            # ======================
            if "Net Receivable \\ Payable" in subset.columns:
                # Calculate total without rounding - always preserve decimals
                total_value = subset["Net Receivable \\ Payable"].apply(lambda x: safe_float_no_round(x) if pd.notna(x) else 0.0).sum()

                total_row = {col: "" for col in subset.columns}
                cols = list(subset.columns)
                net_idx = cols.index("Net Receivable \\ Payable")

                if net_idx > 0:
                    total_row[cols[net_idx - 1]] = "Total"   # left side of sum
                total_row["Net Receivable \\ Payable"] = total_value

                subset = pd.concat([subset, pd.DataFrame([total_row])], ignore_index=True)

            safe_sheet_name = str(key)[:31] or "Blank"  # Excel max 31 chars
            subset.to_excel(writer, sheet_name=safe_sheet_name, index=False)
            
            # Add borders to all cells in the sheet
            worksheet = writer.sheets[safe_sheet_name]
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply borders to all cells with data
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
            
            # Apply comma style formatting to specific columns
            comma_style_columns = [
                "Buy STT",
                "Sell STT",
                "Sell Stamp Duty",
                "Buy Stamp Duty",
                "Buy Payable Amount",
                "Sell Receivable Amount",
                "Net Receivable \\ Payable"
            ]
            
            # Find column indices for comma style columns
            for col_idx, col_name in enumerate(subset.columns, 1):
                if col_name in comma_style_columns:
                    column_letter = get_column_letter(col_idx)
                    # Apply comma format to all data rows (skip header row 1)
                    for row_idx in range(2, len(subset) + 2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
            
            # Auto-adjust column widths based on content
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set minimum width for EXTRA_COLUMNS and adjust based on content
                adjusted_width = max(max_length + 2, 12)  # Add padding and minimum width
                
                # Special handling for EXTRA_COLUMNS to ensure adequate width
                if column_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
                    # Check if this column contains EXTRA_COLUMNS data
                    header_cell = worksheet.cell(row=1, column=column[0].column)
                    if header_cell.value in cons_header.EXTRA_COLUMNS:
                        adjusted_width = max(adjusted_width, 18)  # Ensure extra columns have good width
                
                worksheet.column_dimensions[column_letter].width = adjusted_width