"""
Report Processing Modules
Separated processing logic for different report types
"""

import csv
import os
import re
import math
import pandas as pd
import zipfile
import io
import calendar
import gzip
import shutil
from datetime import datetime, timedelta
import traceback
import glob
import cons_header
from client_position_page import load_passwords
from db_manager import insert_report
from physical_settlement_files import build_dict, segregate_excel_by_column, create_segregated_file_with_summary

import json
import os
import tkinter as tk
from tkinter import messagebox

from decimal import Decimal, ROUND_HALF_UP

from openpyxl import Workbook
import pyzipper
import py7zr

from openpyxl.styles import PatternFill

class BaseProcessor:
    """Base class for all processors"""
    def __init__(self, db_path, log_error_callback):
        self.db_path = db_path
        self.log_error = log_error_callback
    
    def validate_inputs(self, **kwargs):
        """Validate common inputs - override in subclasses"""
        pass
    
    def create_output_directory(self, output_path):
        """Create output directory if it doesn't exist"""
        if not os.path.exists(output_path):
            try:
                os.makedirs(output_path)
            except Exception as e:
                raise Exception(f"Cannot create output directory: {str(e)}")
    
    def handle_file_permission_error(self, file_path, operation="read"):
        """Show user-friendly popup for file permission errors"""
        filename = os.path.basename(file_path)
        
        try:
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            
            message = f"Cannot {operation} the file:\n{filename}\n\nPlease close the file and try again."
            
            messagebox.showerror("File Access Error", message)
            root.destroy()
        except Exception:
            # Fallback if tkinter is not available
            pass
        
        # Return a special value to indicate permission error was handled
        return "PERMISSION_ERROR_HANDLED"


class MonthlyFloatProcessor(BaseProcessor):
    """Processor for Monthly Float Report"""
    
    def process(self, fno_path, mcx_path, output_path):
        """Process FNO and MCX files for monthly float report"""
        try:
            self.validate_inputs(fno_path=fno_path, mcx_path=mcx_path, output_path=output_path)
            self.create_output_directory(output_path)
            
            error_log_path = os.path.join(output_path, "pcm_errors.txt")
            fno_files = glob.glob(os.path.join(fno_path, "*.csv"))
            mcx_files = glob.glob(os.path.join(mcx_path, "*.csv"))

            fno_count = len(fno_files)
            mcx_count = len(mcx_files)
            
            # Merge files
            self._merge_fno_and_mcx(fno_files, mcx_files, output_path, error_log_path)
            
            # Process data
            df_list = []
            for folder in [fno_files, mcx_files]:
                for file in folder:
                    try:
                        temp_df = pd.read_csv(file)
                        temp_df.columns = temp_df.columns.str.strip()
                        df_list.append(temp_df[cons_header.columns_to_keep])
                    except Exception as e:
                        self.log_error(error_log_path, file, e)

            if not df_list:
                raise Exception("No CSV files found or all failed to load.")

            df = pd.concat(df_list, ignore_index=True)

            # Fill missing dates
            df_before_fill = len(df)
            df, messages = self._fill_missing_dates(df, error_log_path)
            df_after_fill = len(df)
            missing_filled = df_after_fill - df_before_fill

            # Generate summary
            summary_data = self._generate_summary(df)
            
            # Write Excel file
            output_file = self._write_excel_output(df, summary_data, output_path)
            
            # Create monthly status log
            self._create_monthly_status_log(messages, output_path)
            
            # Create ZIP and save to database
            self._create_zip_and_save(output_path, fno_path, mcx_path, output_file)
            
            return {
                'fno_count': fno_count,
                'mcx_count': mcx_count,
                'missing_filled': missing_filled,
                'output_file': output_file
            }
            
        except Exception as e:
            self.log_error(output_path, "Monthly Float Processing", e)
            raise e
    
    def validate_inputs(self, fno_path, mcx_path, output_path):
        """Validate inputs for monthly float processing"""
        if not fno_path or not mcx_path or not output_path:
            raise ValueError("Please select all folders before processing.")
    
    def _merge_fno_and_mcx(self, fno_files, mcx_files, output_path, error_log_path):
        """Merge FNO and MCX data"""
        try:
            all_dataframes = []

            for file in fno_files:
                try:
                    df = pd.read_csv(file)
                    df['Source'] = 'FNO'
                    all_dataframes.append(df)
                except Exception as e:
                    self.log_error(error_log_path, file, e)

            for file in mcx_files:
                try:
                    df = pd.read_csv(file)
                    df['Source'] = 'MCX'
                    all_dataframes.append(df)
                except Exception as e:
                    self.log_error(error_log_path, file, e)

            if not all_dataframes:
                raise Exception(f"No valid CSV files found in FNO or MCX.\nSee log: {error_log_path}")

            merged_df = pd.concat(all_dataframes, ignore_index=True)
            output_file = os.path.join(output_path, "merged_fno_mcx_data.xlsx")
            merged_df.to_excel(output_file, index=False)

        except Exception as e:
            self.log_error(error_log_path, "merge_fno_and_mcx", e)
            raise Exception(f"A fatal error occurred.\nSee log: {error_log_path}")
    
    def _fill_missing_dates(self, df, error_log_path):
        """Fill missing dates by duplicating previous day's data"""
        try:
            df[cons_header.DATE] = pd.to_datetime(df[cons_header.DATE], dayfirst=True)
            df["YEAR"] = df[cons_header.DATE].dt.year
            df["MONTH"] = df[cons_header.DATE].dt.month

            filled_data = []
            status_messages = []

            for cp_code, cp_group in df.groupby(cons_header.CP_CODE, dropna=False):
                for (year, month), month_group in cp_group.groupby(["YEAR", "MONTH"]):
                    try:
                        year = int(year)
                        month = int(month)

                        days_in_month = calendar.monthrange(year, month)[1]
                        month_name = calendar.month_name[month]

                        all_days = pd.date_range(start=f"{year}-{month:02d}-01", periods=days_in_month)

                        existing_dates = set(month_group[cons_header.DATE])
                        missing_dates = [d for d in all_days if d not in existing_dates]

                        msg = ""
                        cp_code_display = "blankcpcode" if cp_code == "" or str(cp_code).lower() == "nan" else cp_code
                        if missing_dates:
                            missing_day_nums = ", ".join(str(d.day) for d in missing_dates)
                            msg = f"[INFO] {cp_code_display} → {month_name} {year}: Missing {len(missing_dates)} day(s) → Days: {missing_day_nums}"
                        else:
                            msg = f"[SUCCESS] {cp_code_display} → {month_name} {year}: ✅ All {days_in_month} days present."

                        status_messages.append(msg)
                        filled_month = month_group.copy()

                        for date in missing_dates:
                            prev_data = filled_month[filled_month[cons_header.DATE] < date]
                            if prev_data.empty:
                                continue
                            last_row = prev_data.sort_values(cons_header.DATE).iloc[-1].copy()
                            last_row[cons_header.DATE] = date
                            filled_month = pd.concat([filled_month, pd.DataFrame([last_row])])

                        filled_data.append(filled_month.sort_values(cons_header.DATE))

                    except Exception as e:
                        if error_log_path:
                            self.log_error(error_log_path, f"{cp_code} - {year}-{month:02d}", e)
                        return None, []

            final_df = pd.concat(filled_data, ignore_index=True)
            final_df.drop(columns=["YEAR", "MONTH"], inplace=True)
            return final_df, status_messages

        except Exception as e:
            if error_log_path:
                self.log_error(error_log_path, "fill_missing_dates", e)
            raise e
    
    def _generate_summary(self, df):
        """Generate summary data for the report"""
        summary_data = []
        df[cons_header.CP_CODE] = df[cons_header.CP_CODE].fillna("").astype(str).replace("nan", "").str.strip()
        cp_groups = list(df.groupby(cons_header.CP_CODE, dropna=False))
        
        for cp_code, group in cp_groups:
            cp_code_display = "blankcpcode" if cp_code == "" else cp_code
            total_val = group[cons_header.FINANCIAL_LEDGER_BALANCE].sum()
            avg_val = group[cons_header.FINANCIAL_LEDGER_BALANCE].mean()
            group[cons_header.CP_CODE] = "" if cp_code == "" else cp_code
            summary_data.append({
                "CP Code": cp_code_display,
                "Total": total_val,
                "Average": avg_val
            })
        
        return summary_data
    
    def _write_excel_output(self, df, summary_data, output_path):
        """Write Excel output file"""
        output_file = os.path.join(output_path, "cp_code_separate_sheets.xlsx")
        with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
            summary_df = pd.DataFrame(summary_data, columns=["CP Code", "Total", "Average"])
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

            df[cons_header.CP_CODE] = df[cons_header.CP_CODE].fillna("").astype(str).replace("nan", "").str.strip()
            cp_groups = list(df.groupby(cons_header.CP_CODE, dropna=False))
            
            for cp_code, group in cp_groups:
                sheet_name = "blankcpcode" if cp_code == "" else cp_code[:31]
                group[cons_header.CP_CODE] = "" if cp_code == "" else cp_code
                group.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return output_file
    
    def _create_monthly_status_log(self, messages, output_path):
        """Create monthly status log file"""
        monthly_log_path = os.path.join(output_path, "monthly_status.txt")
        user_friendly_header = "ℹ️ Monthly Status: Missing dates have been filled automatically. Please check the summary below.\n\n"
        full_message = user_friendly_header + "\n".join(messages)

        with open(monthly_log_path, "w", encoding="utf-8") as f:
            f.write(full_message)
    
    def _create_zip_and_save(self, output_path, fno_path, mcx_path, output_file):
        """Create ZIP file and save to database"""
        # Create ZIP
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(output_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, output_path)
                    zipf.write(file_path, arcname)
        zip_blob = zip_buffer.getvalue()

        # Insert into database
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        insert_report(self.db_path, report_type=cons_header.NSE_AND_MCX, 
                     created_at=timestamp, modified_at=timestamp, report_blob=zip_blob)


class NMASSAllocationProcessor(BaseProcessor):
    """Processor for NMASS Allocation Report"""
    
    def process(self, date, sheet, input1_path, input2_path, output_path):
        """Process NMASS allocation files"""
        try:
            self.validate_inputs(date=date, input1_path=input1_path, 
                               input2_path=input2_path, output_path=output_path)
            self.create_output_directory(output_path)
            
            # Process the files
            result = self._process_ledger_files(input1_path, input2_path, date, sheet, output_path)
            
            if result:
                return f"Ledger processed successfully with {result['record_count']} records."
            else:
                raise Exception("Failed to process ledger files.")
                
        except Exception as e:
            self.log_error(output_path, "NMASS Allocation Processing", e)
            raise e
    
    def validate_inputs(self, date, input1_path, input2_path, output_path):
        """Validate inputs for NMASS allocation processing"""
        if date == "DD/MM/YYYY" or not date.strip():
            raise ValueError("Please select a valid date.")
            
        if not input1_path.strip() or not input2_path.strip():
            raise ValueError("Please select both attachment files.")
        
        if not output_path.strip():
            raise ValueError("Please select an output folder for the ledger.")
        
        if not os.path.exists(input1_path):
            raise ValueError(f"System file not found:\n{input1_path}")
            
        if not os.path.exists(input2_path):
            raise ValueError(f"Manual file not found:\n{input2_path}")
    
    def _read_file(self, file_path, selected_sheet=None, **kwargs):
        """Read file with appropriate method based on extension"""
        ext = os.path.splitext(file_path)[1].lower()
        
        try:
            if ext == ".csv":
                df = pd.read_csv(file_path, **kwargs)
            elif ext in [".xls", ".xlsx"]:
                df = pd.read_excel(file_path, sheet_name=selected_sheet or 0, **kwargs)
            else:
                raise ValueError(f"Unsupported file type: {ext}")
            
            # Drop rows where all columns are NaN
            df = df.dropna(how='all')
            return df
        except PermissionError:
            self.handle_file_permission_error(file_path, "read")
        except Exception as e:
            if "Permission denied" in str(e) or "being used by another process" in str(e):
                self.handle_file_permission_error(file_path, "read")
            else:
                raise e
    
    def _get_next_file_path(self, output_path, base_name, dt):
        """Generate the next available file path by incrementing T000X"""
        import re
        pattern = re.compile(rf"{re.escape(base_name)}_ALLOC_{dt}\.T(\d+)$")

        max_num = 0
        for fname in os.listdir(output_path):
            match = pattern.match(fname)
            if match:
                num = int(match.group(1))
                if num > max_num:
                    max_num = num

        next_num = max_num + 1
        filename = f"{base_name}_ALLOC_{dt}.T{next_num:04d}"
        return os.path.normpath(os.path.join(output_path, filename))
    
    def _build_segment_line(self, date, segment, member_code, cp_code, c_value, margin_value, status):
        """Build segment line for the report"""
        return f"{date},{segment},{member_code},,{cp_code},,{c_value},{margin_value},,,,,,,{status}"
    
    def _process_ledger_files(self, file1_path, file2_path, selected_date, selected_sheet, output_path):
        """Process ledger files and perform calculations"""
        try:
            df1 = self._read_file(file1_path)
            ext = os.path.splitext(file2_path)[1].lower()

            if ext == ".csv":
                df2 = self._read_file(file2_path, header=9, usecols="B:K")
            elif ext in [".xls", ".xlsx"]:
                df2 = self._read_file(file2_path, header=9, usecols=[cons_header.CLIENT_CODE, cons_header.FO_MARGIN])
            else:
                raise ValueError(f"Unsupported file type: {ext}")
            
            df1[cons_header.TM_CP_CODE] = df1[cons_header.TM_CP_CODE].astype(str)
            df2[cons_header.CLIENT_CODE] = df2[cons_header.CLIENT_CODE].astype(str)

            data_dict = {
                "system": dict(zip(df1[cons_header.TM_CP_CODE], df1[cons_header.CASH_COLLECTED])),
                "manual": dict(zip(df2[cons_header.CLIENT_CODE], df2[cons_header.FO_MARGIN]))
            }
            
            dict1 = data_dict["system"]
            dict2 = data_dict["manual"]

            formatted_date = datetime.strptime(selected_date, "%d/%m/%Y").strftime("%d-%b-%Y").upper()
            dt = datetime.strptime(selected_date, "%d/%m/%Y").strftime("%d%m%Y")
            processed_lines = set()

            file_path = self._get_next_file_path(output_path, cons_header.NSE_MEMBER_CODE, dt)
            lines_to_write = []

            # Process existing keys
            for key in dict1:
                if key.lower() == "nan":
                    continue
                if key in dict2:
                    difference = dict2[key] - dict1[key]
                    
                    if difference > 0:
                        status = "U"
                    elif difference < 0:
                        status = "D"
                    else:
                        continue

                    line_fo = self._build_segment_line(formatted_date, cons_header.SEGMENTS[selected_sheet], 
                                                     cons_header.NSE_MEMBER_CODE, key, cons_header.C, dict2[key], status)

                    if line_fo not in processed_lines:
                        lines_to_write.append(line_fo)
                        processed_lines.add(line_fo)
            
            # Process keys in dict2 but not in dict1
            for key in dict2:
                if key.lower() == "nan":
                    continue
                if key not in dict1:
                    if float(dict2[key]) == 0:
                        continue
                    status = "U"
                    line_fo = self._build_segment_line(formatted_date, cons_header.SEGMENTS[selected_sheet], 
                                                     cons_header.NSE_MEMBER_CODE, key, cons_header.C, dict2[key], status)
                    if line_fo not in processed_lines:
                        lines_to_write.append(line_fo)
                        processed_lines.add(line_fo)
            
            # Sort so that 'D' comes before 'U'
            sorted_lines = sorted(lines_to_write, key=lambda x: x.strip().split(",")[-1])
            for i in sorted_lines:
                if i.split(",")[4] == '90072':
                    sorted_lines.remove(i)

            # Write lines into report file
            with open(file_path, "w") as f:
                if lines_to_write:
                    f.write("\n".join(sorted_lines))
                else:
                    f.write("")

            # Create ZIP and save to database
            self._create_zip_and_save(file1_path, file2_path, file_path, output_path, dt)
            
            return {"record_count": len(sorted_lines)}

        except Exception as e:
            self.log_error(output_path, "Error in process_ledger_files", e)
            return None
    
    def _create_zip_and_save(self, file1_path, file2_path, output_file, output_path, dt):
        """Create ZIP file and save to database"""
        # Create ZIP
        zip_filename = f"{cons_header.NSE_MEMBER_CODE}_REPORT_{dt}.zip"
        zip_path = os.path.join(output_path, zip_filename)
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            zipf.write(file1_path, os.path.basename(file1_path))
            zipf.write(file2_path, os.path.basename(file2_path))
            zipf.write(output_file, os.path.basename(output_file))

        # Read ZIP as binary and insert into DB
        with open(zip_path, 'rb') as f:
            zip_blob = f.read()

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        insert_report(self.db_path, report_type=cons_header.LEDGER, 
                     created_at=timestamp, modified_at=timestamp, report_blob=zip_blob)


class ObligationSettlementProcessor(BaseProcessor):
    """Processor for Obligation Settlement"""
    
    def process(self, obligation_path, stt_path, stamp_duty_path, output_path):
        """Process obligation settlement files"""
        try:
            self.validate_inputs(obligation_path=obligation_path, stt_path=stt_path, 
                               stamp_duty_path=stamp_duty_path, output_path=output_path)
            self.create_output_directory(output_path)
            
            # Build dictionaries
            stt_dict = build_dict(
                file_path=stt_path,
                key_cols=["BrkrOrCtdnPtcptId","TckrSymb", "FinInstrmId"],
                value_cols={
                    "Buy STT": "BuyDelvryTtlTaxs",
                    "Sell STT": "SellDelvryTtlTaxs"
                },
                filter_col="RptHdr",
                filter_val=40,
                no_round_cols=["Buy STT", "Sell STT"]
            )

            stamp_duty_dict = build_dict(file_path=stamp_duty_path,
                key_cols=["BrkrOrCtdnPtcptId","TckrSymb", "FinInstrmId"],
                value_cols={
                    "Sell Stamp Duty": "BuyOthrThanDlvryStmpDty",
                    "Buy Stamp Duty": "BuyDlvryStmpDty"
                },
                filter_col="RptHdr",
                filter_val=40,
                no_round_cols=["Buy Stamp Duty", "Sell Stamp Duty"]
            )
            
            output_file = os.path.join(output_path, "Physical_Settlement_Report.xlsx")

            # Segregate obligation and update Buy/Sell STT from dictionary
            segregate_excel_by_column(
                excel_path=obligation_path,
                output_path=output_file,
                column_name="BrkrOrCtdnPtcptId",
                custom_header=cons_header.OBLIGATION_HEADER,
                update_dicts=[stt_dict, stamp_duty_dict]
            )
            
            # Create segregated files with summary in ZIP
            segregated_output_zip = os.path.join(output_path, "Physical_Settlement_Segregated_Report.zip")
            create_segregated_file_with_summary(
                obligation_path=obligation_path,
                stt_path=stt_path,
                stamp_duty_path=stamp_duty_path,
                output_path=segregated_output_zip,
                column_name="BrkrOrCtdnPtcptId",
                custom_header=cons_header.OBLIGATION_HEADER,
                update_dicts=[stt_dict, stamp_duty_dict]
            )

            # Create ZIP and save to database
            self._create_zip_and_save(obligation_path, stt_path, stamp_duty_path, output_file, output_path)
            
            return f"Physical Settlement processed successfully. Output: {output_file}"

        except Exception as e:
            self.log_error(output_path, "Physical Settlement Processing", e)
            raise e
    
    def validate_inputs(self, obligation_path, stt_path, stamp_duty_path, output_path):
        """Validate inputs for obligation settlement processing"""
        if not obligation_path or not stt_path or not stamp_duty_path or not output_path:
            raise ValueError("Please select all input files and output folder.")
    
    def _create_zip_and_save(self, obligation_path, stt_path, stamp_duty_path, output_file, output_path):
        """Create ZIP file and save to database"""
        # Create ZIP
        dt = datetime.now().strftime("%Y%m%d_%H%M%S")
        zip_filename = f"{cons_header.NSE_MEMBER_CODE}_PHYSICAL_SETTLEMENT_{dt}.zip"
        zip_path = os.path.join(output_path, zip_filename)

        with zipfile.ZipFile(zip_path, 'w') as zipf:
            zipf.write(obligation_path, os.path.basename(obligation_path))
            zipf.write(stt_path, os.path.basename(stt_path))
            zipf.write(stamp_duty_path, os.path.basename(stamp_duty_path))
            zipf.write(output_file, os.path.basename(output_file))

        # Insert ZIP into DB
        with open(zip_path, 'rb') as f:
            zip_blob = f.read()
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        insert_report(self.db_path, report_type="PHYSICAL_SETTLEMENT", 
                     created_at=timestamp, modified_at=timestamp, report_blob=zip_blob)


class SegregationReportProcessor(BaseProcessor):
    """Processor for Segregation Report"""
    
    def process(self, date, cp_pan, cash_collateral_cds, cash_collateral_fno,
                daily_margin_nsecr, daily_margin_nsefno, x_cp_master, f_cp_master,
                collateral_valuation_cds, collateral_valuation_fno,
                sec_pledge, 
                cash_with_ncl, santom_file, extra_records, output_path):
        """Process segregation report files"""
        try:
            self.validate_inputs(date=date, cp_pan=cp_pan, output_path=output_path,
                               cash_collateral_cds=cash_collateral_cds, cash_collateral_fno=cash_collateral_fno,
                               daily_margin_nsecr=daily_margin_nsecr, daily_margin_nsefno=daily_margin_nsefno,
                               x_cp_master=x_cp_master, f_cp_master=f_cp_master,
                               collateral_valuation_cds=collateral_valuation_cds, collateral_valuation_fno=collateral_valuation_fno, 
                              sec_pledge=sec_pledge,
                               cash_with_ncl=cash_with_ncl, santom_file=santom_file, extra_records=extra_records)
            self.create_output_directory(output_path)
            
            # Process the segregation report
            result = self._process_segregation_files(
                date, cp_pan, cash_collateral_cds, cash_collateral_fno,
                daily_margin_nsecr, daily_margin_nsefno, x_cp_master, f_cp_master,
                collateral_valuation_cds, collateral_valuation_fno, 
                sec_pledge, 
                cash_with_ncl, santom_file, extra_records, output_path
            )
            
            return result
                
        except Exception as e:
            self.log_error(output_path, "Segregation Report Processing", e)
            raise e
    
    def validate_inputs(self, date, cp_pan, output_path, **file_paths):
        """Validate inputs for segregation report processing"""
        if date == "DD/MM/YYYY" or not date.strip():
            raise ValueError("Please select a valid date.")
            
        if not cp_pan.strip():
            raise ValueError("Please enter CP PAN.")
        
        # Check required files
        missing_files = []
        for file_name, file_path in file_paths.items():
            if file_name == "cash_with_ncl" or file_name == "santom_file":
                continue
            if not file_path.strip():
                missing_files.append(file_name.replace('_', ' ').title())
        
        if missing_files:
            raise ValueError(f"Please select the following files:\n" + "\n".join(f"- {name}" for name in missing_files))
        
        if not output_path.strip():
            raise ValueError("Please select an output folder.")
        
        # Check if files exist
        for file_name, file_path in file_paths.items():
            
            if file_name == "cash_with_ncl":
                continue

            if file_path and not os.path.exists(file_path):
                raise ValueError(f"File not found: {file_name.replace('_', ' ').title()}\n{file_path}")
    
    def _process_segregation_files(self, date, cp_pan, cash_collateral_cds, cash_collateral_fno,
                                daily_margin_nsecr, daily_margin_nsefno, x_cp_master, f_cp_master,
                                collateral_valuation_cds, collateral_valuation_fno, 
                                sec_pledge, 
                                cash_with_ncl, santom_file, extra_records, output_path):
        """Process all segregation files and generate the final report"""
        try:
            # Import segregation functions
            from segregation import read_file, write_file, build_cp_lookup
            from CONSTANT_SEGREGATION import segregation_headers, A, B, C, D, E, F, G, H, I, J, K, L, O, P, AD, AV, AG, AW, AH, AX, BB, BD, BF, AT
            
            # Format date for output
            formatted_date = datetime.strptime(date, "%d/%m/%Y").strftime("%d-%m-%Y")
            
            # Read CP Master files
            try:
                df_fo = read_file(f_cp_master)
                cp_codes_fo = df_fo["CP Code"].tolist()
                pan_fo = df_fo["PAN Number"].tolist()
            except Exception as e:
                if "permission error" in str(e).lower():
                    return self.handle_file_permission_error(f_cp_master, "read")
                else:
                    raise Exception(f"❌ Error reading F_CPMaster_data file:\n\nPlease check if the correct F_CPMaster_data file is attached.\n\nTechnical details: {str(e)}")
            
            try:
                df_cd = read_file(x_cp_master)
                cp_codes_cd = df_cd["CP Code"].tolist()
                pan_cd = df_cd["PAN Number"].tolist()
            except Exception as e:
                if "permission error" in str(e).lower():
                    return self.handle_file_permission_error(x_cp_master, "read")
                else:
                    raise Exception(f"❌ Error reading X_CPMaster_data file:\n\nPlease check if the correct X_CPMaster_data file is attached.\n\nTechnical details: {str(e)}")
            
            # Read Cash Collateral files
            try:
                df_cash_cds = read_file(cash_collateral_cds, header_row=9, usecols="B:I")
                cd_collateral_lookup = dict(zip(df_cash_cds["ClientCode"], df_cash_cds["TotalCollateral"]))
            except Exception as e:
                if "permission error" in str(e).lower():
                    return self.handle_file_permission_error(cash_collateral_cds, "read")
                else:
                    raise Exception(f"❌ Error reading CashCollateral_CDS file:\n\nPlease check if the correct CashCollateral_CDS file is attached.\n\nTechnical details: {str(e)}")
            
            try:
                df_cash_fno = read_file(cash_collateral_fno, header_row=9, usecols="B:I")
                fo_collateral_lookup = dict(zip(df_cash_fno["ClientCode"], df_cash_fno["TotalCollateral"]))
            except Exception as e:
                if "permission error" in str(e).lower():
                    return self.handle_file_permission_error(cash_collateral_fno, "read")
                else:
                    raise Exception(f"❌ Error reading CashCollateral_FNO file:\n\nPlease check if the correct CashCollateral_FNO file is attached.\n\nTechnical details: {str(e)}")
            
            # Read Daily Margin files
            try:
                df_margin_cds = read_file(daily_margin_nsecr, header_row=9, usecols="B:T")
                cd_daily_margin_lookup = dict(zip(df_margin_cds["ClientCode"], df_margin_cds["Funds"]))
            except Exception as e:
                raise Exception(f"❌ Error reading Daily Margin Report NSECR file:\n\nPlease check if the correct Daily Margin Report NSECR file is attached.\n\nTechnical details: {str(e)}")
            
            try:
                df_margin_fno = read_file(daily_margin_nsefno, header_row=9, usecols="B:T")
                fo_daily_margin_lookup = dict(zip(df_margin_fno["ClientCode"], df_margin_fno["Funds"]))
            except Exception as e:
                raise Exception(f"❌ Error reading Daily Margin Report NSEFNO file:\n\nPlease check if the correct Daily Margin Report NSEFNO file is attached.\n\nTechnical details: {str(e)}")
            
            # Read Collateral Valuation Report CD
            try:
                df_valuation_cd = read_file(collateral_valuation_cds, header_row=9, usecols="B:H")
                cd_collateral_valuation_lookup = {}

                for _, row in df_valuation_cd.iterrows():
                    client_code = row["ClientCode"]
                    cash_eq = row["CashEquivalent"]
                    non_cash = row["NonCash"]
                    
                    if client_code in cd_collateral_valuation_lookup:
                        cd_collateral_valuation_lookup[client_code]["CashEquivalent"] = cash_eq
                        cd_collateral_valuation_lookup[client_code]["NonCash"] = non_cash
                    else:
                        cd_collateral_valuation_lookup[client_code] = {
                            "CashEquivalent": cash_eq,
                            "NonCash": non_cash
                        }
            except Exception as e:
                raise Exception(f"❌ Error reading Collateral Valuation Report CDS file:\n\nPlease check if the correct Collateral Valuation Report CDS file is attached.\n\nTechnical details: {str(e)}")

            # Read Collateral Valuation Report FO
            try:
                df_valuation_fo = read_file(collateral_valuation_fno, header_row=9, usecols="B:H")
                fo_collateral_valuation_lookup = {}
                
                for _, row in df_valuation_fo.iterrows():
                    client_code = row["ClientCode"]
                    cash_eq = row["CashEquivalent"]
                    non_cash = row["NonCash"]
                    
                    if client_code in fo_collateral_valuation_lookup:
                        fo_collateral_valuation_lookup[client_code]["CashEquivalent"] = cash_eq
                        fo_collateral_valuation_lookup[client_code]["NonCash"] = non_cash
                    else:
                        fo_collateral_valuation_lookup[client_code] = {
                            "CashEquivalent": cash_eq,
                            "NonCash": non_cash
                        }
            except Exception as e:
                raise Exception(f"❌ Error reading Collateral Valuation Report FNO file:\n\nPlease check if the correct Collateral Valuation Report FNO file is attached.\n\nTechnical details: {str(e)}")
            
            # Process Security Pledge file
            try:
                sec_pledge_cp_lookup = self._process_security_pledge_file(sec_pledge)
            except Exception as e:
                if "permission error" in str(e).lower():
                    return self.handle_file_permission_error(sec_pledge, "read")
                else:
                    raise Exception(f"❌ Error reading Security Pledge file:\n\nPlease check if the correct Security Pledge file is attached.\n\nTechnical details: {str(e)}")
            
            # Generate report data
            data = self._generate_report_data(
                formatted_date, cp_pan, cp_codes_fo, pan_fo, cp_codes_cd, pan_cd,
                fo_collateral_lookup, fo_daily_margin_lookup, cd_collateral_lookup, 
                cd_daily_margin_lookup, cd_collateral_valuation_lookup,fo_collateral_valuation_lookup, sec_pledge_cp_lookup
            )
            # Load master records using simple dynamic function
            av_records, at_records = self._get_master_records() # Get Both AV and AT Records (Default):
            # 2. Get Only AV or AT Records:
            # av_records = self._get_master_records(av=True) at_records = self._get_master_records(at=True)
            # all_records = self._get_master_records(all_records=True)
            
            for data_record in data:
                try:
                    cp_key = str(data_record.get(D, "")).strip()
                    seg_key = str(data_record.get(H, "")).strip()
                    if not (cp_key and seg_key):
                        continue

                    for av_record in av_records or []:
                        av_cp = (av_record.get(D) or "").strip()
                        av_seg = (av_record.get(H) or "").strip()
                        if av_cp == cp_key and av_seg == seg_key:
                            av_val_raw = av_record.get("av_value") if "av_value" in av_record else av_record.get(AV)
                            if av_val_raw not in (None, ""):
                                try:
                                    data_record[AV] = float(av_val_raw)
                                except Exception:
                                    pass
                            break  # stop at first match
                except Exception:
                    continue

            # Process extra records first
            extra_records_data = []
            if extra_records:
                try:
                    extra_records_df = read_file(extra_records)
                    for _, row in extra_records_df.iterrows():
                        record = {}
                        for col in extra_records_df.columns:
                            val = row[col]

                            if col == A:
                                # Use formatted_date from frontend instead of parsing from data
                                val = formatted_date
                                
                            record[col] = val
                        
                        # Custom logic
                        if str(row.get(G, "")).strip() == "P" and str(row.get(H, "")).strip() == "FO":
                            # Lookup in AV_Records
                            for av_record in av_records:
                                if (
                                    av_record.get(G) == "P" and
                                    av_record.get(H) == "FO"
                                ):
                                    record[AV] = av_record["av_value"]
                                    break  # stop at first match


                        extra_records_data.append(record)
                except Exception as e:
                    raise Exception(f"❌ Error reading Extra_Records_File:\n\nPlease check if the correct Extra_Records_File is attached.\n\nTechnical details: {str(e)}")
            
            # Filter main data and add extra records in correct position
            data = self._segregation_data_filter(data, segregation_headers=segregation_headers[9:], extra_records=extra_records_data)
            
            # Loop through data (list of dictionaries) and apply AT records logic
            for i, data_record in enumerate(data):
                # Loop through AT records to find matches
                for at_record in at_records:
                    at_cp_code = at_record.get(D, '')
                    at_segment = at_record.get(H, '')
                    at_value = float(at_record.get("at_value", 0))
                    
                    # Check if current data record matches AT record criteria
                    if (data_record.get(D, '') == at_cp_code and 
                        data_record.get(H, '') == at_segment):
                        
                        # Apply AT logic to this data record

                        data_record[AV] = data_record[AD] - at_value
                        data_record[AT] = at_value

                        break  # Stop at first match
            
            if santom_file:
                try:
                    santom_df = read_file(santom_file)
                    data = self._santom_file_working(data, cash_with_ncl, santom_df)
                except Exception as e:
                    raise Exception(f"❌ Error reading SANTOM_FILE:\n\nPlease check if the correct SANTOM_FILE is attached.\n\nTechnical details: {str(e)}")

            # Write output file
            output_file = os.path.join(output_path, f"{cp_pan}_{formatted_date.replace('-', '')}_01.csv")
            write_file(output_file, data=data, header=segregation_headers)

            # Also save the CSV into a ZIP with the same base name
            csv_zip_path = None
            try:
                csv_base_name = os.path.splitext(os.path.basename(output_file))[0]
                csv_zip_path = os.path.join(output_path, f"{csv_base_name}.zip")
                with zipfile.ZipFile(csv_zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zipf:
                    zipf.write(output_file, os.path.basename(output_file))
            except Exception as zip_err:
                # Non-fatal: proceed even if CSV zip creation fails
                pass

            # Create ZIP and save to database
            self._create_zip_and_save(
                cash_collateral_cds, cash_collateral_fno, daily_margin_nsecr, daily_margin_nsefno,
                x_cp_master, f_cp_master, collateral_valuation_cds, collateral_valuation_fno, 
                sec_pledge,
                output_file, output_path
            )
            
            return f"Segregation report generated successfully with {len(data)} records."
            
        except Exception as e:
            self.log_error(output_path, "Error in process_segregation_files", e)
            return None
    
    def _process_security_pledge_file(self, sec_pledge):
        """Process security pledge file"""
        from segregation import read_file, GSEC_HEADER, SRNO,CPCODE,SEGMENT,ISIN,PLEDGE_TYPE,SEC_NAME,MATURITY_DATE,PRICE,QUANTITY,MKT_VALUE,HAIRCUT,HAIRCUT_VALUE,POST_HAIRCUT, D, H

        gsec_df = read_file(
                        sec_pledge,
                        header_row=0,
                        sheet_name="Valuation_G-Sec"
                    )
        gsec_df.columns = gsec_df.columns.str.strip()  # removes leading/trailing spaces

        _sec_pledge_lookup = {}

        for idx, row in gsec_df.iterrows():
            cp_code = str(row[CPCODE]).strip()
            segment = str(row[SEGMENT]).strip()
            pledge_type = str(row[PLEDGE_TYPE]).strip()
            post_haircut = float(row[POST_HAIRCUT]) if pd.notna(row[POST_HAIRCUT]) else 0.0

            # Only include FNO + E-Kuber rows
            if segment == "FNO" and pledge_type == "E-Kuber":
                if cp_code not in _sec_pledge_lookup:
                    _sec_pledge_lookup[cp_code] = {
                        H: segment,
                        D: cp_code,
                        "post_haircut": 0.0
                    }
                # Add to total
                _sec_pledge_lookup[cp_code]["post_haircut"] += post_haircut
        
        for cp_code in _sec_pledge_lookup:
            value = _sec_pledge_lookup[cp_code]["post_haircut"]
            dec_value = Decimal(str(value))  # convert float → Decimal
            _sec_pledge_lookup[cp_code]["post_haircut"] = dec_value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            # _sec_pledge_lookup[cp_code]["post_haircut"] = round(
            #     _sec_pledge_lookup[cp_code]["post_haircut"], 2
            # )
        
        # breakpoint()
        return _sec_pledge_lookup

        # sec plesge file logic is below
        if False:
            with open(sec_pledge, newline='', encoding="utf-8", errors="ignore") as f:
                reader = csv.reader(f)
                rows = list(reader)

            # Step 1: Find where "GSEC" occurs in first column
            header_row = None
            for idx, row in enumerate(rows):
                if row and row[0].strip().upper() == "GSEC":
                    # Prefer next line if it looks like header
                    header_row = idx + 1
                    break

            if header_row is None:
                raise ValueError("'GSEC' not found in file")

            # Step 2: Extract headers and data
            headers = [col.strip() for col in rows[header_row]]
            data_rows = rows[header_row + 1:]

            # Step 3: Build lookup dictionary
            try:
                col_client = headers.index("Client/CP code")
                col_isin = headers.index("ISIN")
                col_gross = headers.index("GROSS VALUE")
                col_haircut = headers.index("HAIRCUT")
            except ValueError as e:
                raise ValueError(f"❌ Expected column missing: {e}")

            for row in data_rows:
                if len(row) <= max(col_client, col_isin, col_gross, col_haircut):
                    continue  # skip short/incomplete rows

                client_code = row[col_client].strip()
                isin = row[col_isin].strip()
                gross_value = row[col_gross].strip()
                haircut = row[col_haircut].strip()

                if not client_code or not isin:
                    continue

                key = f"{client_code}-{isin}"
                _sec_pledge_lookup[key] = {
                    "GROSS VALUE": gross_value,
                    "HAIRCUT": haircut,
                }

            from segregation import build_cp_lookup
            return build_cp_lookup(_sec_pledge_lookup)
    
    def _generate_report_data(self, formatted_date, cp_pan, 
                              cp_codes_fo, pan_fo, 
                              cp_codes_cd, pan_cd,
                              fo_collateral_lookup, fo_daily_margin_lookup, 
                              cd_collateral_lookup, cd_daily_margin_lookup, 
                              cd_collateral_valuation_lookup, fo_collateral_valuation_lookup, 
                              sec_pledge_cp_lookup):
        """Generate report data for both FO and CD segments"""
        from CONSTANT_SEGREGATION import A, B, C, D, E, F, G, H, I, J, K, L, O, P, AD, AV, AG, AW, AH, AX, BB, BD, BF
        
        data = []
        account_type = "C"
        
        # Process FO data
        for cp, pan_no in zip(cp_codes_fo, pan_fo):
            cv_lookup = fo_collateral_valuation_lookup.get(cp, {"CashEquivalent": 0, "NonCash": 0})
            row = {
                A: formatted_date,
                B: cp_pan,
                C: cp_pan,
                D: cp,
                E: pan_no,
                F: "",  # Client PAN
                G: account_type,
                H: "FO",
                I: "",  # UCC Code
                J: fo_collateral_lookup.get(cp, 0),
                K: fo_daily_margin_lookup.get(cp, 0),
                L: fo_daily_margin_lookup.get(cp, 0),
                O: cv_lookup["CashEquivalent"],
                P: cv_lookup["NonCash"],
                # BB: cv_lookup["CashEquivalent"],
                # BD: cv_lookup["CashEquivalent"],
                # BF: cv_lookup["CashEquivalent"]
            }
            
            # Duplicate values in other columns
            row[AD] = row[K]
            row[AV] = row[K]
            row[AG] = row[O]
            row[AW] = row[O]
            row[AH] = row[P]
            row[AX] = row[P]

            # Apply post_haircut only for FO
            if sec_pledge_cp_lookup:
                pledge_info = sec_pledge_cp_lookup.get(cp)                
                if pledge_info and pledge_info.get(H) == "FNO":
                    val = pledge_info.get("post_haircut", 0.0)
                    row[BB] = val
                    row[BD] = val
                    row[BF] = val

            data.append(row)
        
        # Process CD data
        for cp, pan_no in zip(cp_codes_cd, pan_cd):
            cv_lookup = cd_collateral_valuation_lookup.get(cp, {"CashEquivalent": 0, "NonCash": 0})
            row = {
                A: formatted_date,
                B: cp_pan,
                C: cp_pan,
                D: cp,
                E: pan_no,
                F: "",  # Client PAN
                G: account_type,
                H: "CD",
                I: "",  # UCC Code
                J: cd_collateral_lookup.get(cp, 0),
                K: cd_daily_margin_lookup.get(cp, 0),
                L: cd_daily_margin_lookup.get(cp, 0),
                O: cv_lookup["CashEquivalent"],
                P: cv_lookup["NonCash"],
                # BB: cv_lookup["CashEquivalent"],
                # BD: cv_lookup["CashEquivalent"],
                # BF: cv_lookup["CashEquivalent"]
            }
            
            # Duplicate values in other columns
            row[AD] = row[K]
            row[AV] = row[K]
            row[AG] = row[O]
            row[AW] = row[O]
            row[AH] = row[P]
            row[AX] = row[P]

            # Apply post_haircut only for FO
            if sec_pledge_cp_lookup:
                pledge_info = sec_pledge_cp_lookup.get(cp)
                if pledge_info and pledge_info.get(H) == "CDS":
                    val = pledge_info.get("post_haircut", 0.0)
                    row[BB] = val
                    row[BD] = val
                    row[BF] = val
            
            data.append(row)
        
        return data
    
    def _segregation_data_filter(self, data, segregation_headers, cp_code_col="CP Code", seg_col="Segment Indicator", extra_records=None):
        """
        Filter and normalize segregation data:
        1. Replace blank/NA values with 0 for segregation_headers
        2. Sort by CP Code
        3. Sort by Segment Indicator
        4. Move all-zero rows to the end
        5. Set AZ and BL to "NA" for all records
        
        Args:
            data (list[dict]): list of row dictionaries
            segregation_headers (list[str]): expected headers for segregation
            cp_code_col (str): CP Code column name
            seg_col (str): Segment Indicator column name
        
        Returns:
            list[dict]: filtered and sorted data
        """
        from CONSTANT_SEGREGATION import AZ, BL
        # Step 1: Normalize data - replace blank/NA values with 0 for segregation_headers
        normalized = []
        for row in data:
            new_row = {}
            
            # Process segregation_headers columns
            for col in segregation_headers:
                val = row.get(col, 0)  # default if missing
                if val is None or (isinstance(val, str) and (val.strip() == "" or val.strip().upper() == "NA")):
                    val = 0
                new_row[col] = val

            # Copy other columns as-is
            for key, val in row.items():
                if key not in segregation_headers:
                    new_row[key] = val

            normalized.append(new_row)
        
        # Step 2 & 3: Sort only by Segment Indicator A to Z
        seg_sorted = sorted(normalized, key=lambda x: str(x.get(seg_col, "")).strip().upper())
        
        # Step 4: Move all-zero rows to the end
        def is_all_zero(row):
            """Check if all segregation_headers columns have zero values"""
            return all(
                (v == 0 or v == "0" or str(v).strip() == "0") 
                for k, v in row.items() 
                if k in segregation_headers
            )

        # Separate zero and non-zero rows
        zero_rows = []
        non_zero_rows = []
        
        for row in seg_sorted:
            if is_all_zero(row):
                zero_rows.append(row)
            else:
                non_zero_rows.append(row)
        
        # Combine: non-zero rows first, then extra records, then zero rows
        if extra_records:
            final_data = non_zero_rows + extra_records + zero_rows
        else:
            final_data = non_zero_rows + zero_rows

        # Set AZ and BL to "NA" for all records centrally
        for row in final_data:
            row[AZ] = "NA"
            row[BL] = "NA"

        return final_data

    def _santom_file_working(self, data, cash_with_ncl, santom_df):
        from CONSTANT_SEGREGATION import segregation_headers, A, B, C, D, E, F, G, H, I, J, K, L, O, P, AD, AV, AG, AW, AH, AX, BB, BD, BF, AT, AZ, BL

        for _, row in santom_df.iterrows():
            record = {}
            
            # Copy all columns from santom_df to record
            for col in santom_df.columns:
                record[col] = row[col]
                # try:
                #     float_val = float(row[col])
                #     if float_val.is_integer():
                #         record[col] = int(float_val)  # 123.0 → 123
                #     else:
                #         record[col] = float_val       # 123.45 → 123.45
                # except:
                #     record[col] = row[col]            # Keep as string if conversion fails

            # Check if Account Type is "P" and perform special processing
            if G in santom_df.columns and row[G] == "P":
                # For Account Type "P", calculate balance and assign to AV
                if AD in santom_df.columns:
                    try:
                        balance = float(row[AD]) - float(cash_with_ncl or 0)
                        record[AV] = balance
                        record[AW] = row[AG]
                        record[AX] = row[AH]
                        record[AT] = cash_with_ncl
                    except (ValueError, TypeError):
                        record[AV] = 0
                        record[AW] = row[AG]
                        record[AX] = row[AH]
                        record[AT] = cash_with_ncl
            else:
                # For other account types, copy data from specific columns
                if AD in santom_df.columns:
                    record[AV] = row[AD]
                if AG in santom_df.columns:
                    record[AW] = row[AG]
                if AH in santom_df.columns:
                    record[AX] = row[AH]

            data.append(record)

            for row in data:
                if not row.get(AZ) or pd.isna(row.get(AZ)):  # Checks if AZ is blank or missing
                    row[AZ] = "NA"
                if not row.get(BL) or pd.isna(row.get(BL)):  # Checks if BL is blank or missing
                    row[BL] = "NA"

        return data

    def _create_zip_and_save(self, cash_collateral_cds, cash_collateral_fno, daily_margin_nsecr, 
                           daily_margin_nsefno, x_cp_master, f_cp_master, cd_collateral_valuation_lookup, fo_collateral_valuation_lookup,
                             sec_pledge, 
                             output_file, output_path):
        """Create ZIP file and save to database"""
        # Create ZIP file
        dt = datetime.now().strftime("%Y%m%d_%H%M%S")
        zip_filename = f"SEGREGATION_REPORT_{dt}.zip"
        zip_path = os.path.join(output_path, zip_filename)
        
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            # Add input files
            input_files = [
                (cash_collateral_cds, "CashCollateral_CDS.xls"),
                (cash_collateral_fno, "CashCollateral_FNO.xls"),
                (daily_margin_nsecr, "Daily_Margin_Report_NSECR.xls"),
                (daily_margin_nsefno, "Daily_Margin_Report_NSEFNO.xls"),
                (x_cp_master, "X_CPMaster_data.xlsx"),
                (f_cp_master, "F_CPMaster_data.xlsx"),
                (cd_collateral_valuation_lookup, "Collateral Valuation Report_cds.xls"),
                (fo_collateral_valuation_lookup, "Collateral Valuation Report_fno.xls"),
                (sec_pledge, "G-Sec Valuation.xlsx")
            ]
            
            for file_path, arcname in input_files:
                zipf.write(file_path, arcname)
            
            # Add output file
            zipf.write(output_file, os.path.basename(output_file))
        
        # Insert into database
        with open(zip_path, 'rb') as f:
            zip_blob = f.read()
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        insert_report(self.db_path, report_type="SEGREGATION_REPORT", 
                     created_at=timestamp, modified_at=timestamp, report_blob=zip_blob)
        
        # Delete only the output CSV file after successful ZIP creation and database save
        try:
            if os.path.exists(output_file):
                os.remove(output_file)
        except Exception as e:
            pass
    
    def _get_master_records(self, av=False, at=False, all_records=False):
        """
        Simple function to get master records based on flags
        
        Args:
            av (bool): Return AV_Records if True
            at (bool): Return AT_Records if True  
            all_records (bool): Return all records combined if True
            
        Returns:
            list: Requested records
        """
        import json
        import os
        
        master_records_json_path = "master_records.json"
        av_records = []
        at_records = []

        if os.path.exists(master_records_json_path):
            try:
                with open(master_records_json_path, 'r') as f:
                    all_master_data = json.load(f)

                av_records = all_master_data.get('AV_Records', [])
                at_records = all_master_data.get('AT_Records', [])


            except Exception as e:
                pass
        
        # Return based on flags
        if av:
            return av_records
        elif at:
            return at_records
        elif all_records:
            return av_records + at_records
        else:
            return av_records, at_records  # Default: return both separately


class ClientPositionProcessor(BaseProcessor):
    """Processor for Client Position Report"""
    
    def process(self, client_position_path, output_path, selected_cp_codes=None, cp_codes_config=None, cash_collateral_path=None):
        """Process client position file"""
        try:
            self.validate_inputs(client_position_path=client_position_path, output_path=output_path,
                               selected_cp_codes=selected_cp_codes)
            self.create_output_directory(output_path)
            
            # Process the client position file
            result = self._process_client_position_file(
                client_position_path,
                output_path,
                selected_cp_codes,
                cp_codes_config,
                cash_collateral_path=cash_collateral_path
            )
            
            if result:
                # Check if result contains a friendly message (info status)
                if isinstance(result, dict) and result.get('status') == 'info':
                    return result['message']
                else:
                    return f"Client position processed successfully with {result['record_count']} records for {result['cp_count']} CP code(s)."
            else:
                raise Exception("Failed to process client position file.")
                
        except Exception as e:
            self.log_error(output_path, "Client Position Processing", e)
            raise e
    
    def create_7z_archive(out_path, base_name, excel_file, password=None):
        archive_file = os.path.join(out_path, f"{base_name}.7z")
        
        with py7zr.SevenZipFile(archive_file, 'w', password=password) as archive:
            archive.write(excel_file, arcname=os.path.basename(excel_file))

    def validate_inputs(self, client_position_path, output_path, selected_cp_codes=None):
        """Validate inputs for client position processing"""
        if not client_position_path.strip():
            raise ValueError("Please select a client position file.")
        
        if not output_path.strip():
            raise ValueError("Please select an output folder.")
        
        if not os.path.exists(client_position_path):
            raise ValueError(f"Client position file not found:\n{client_position_path}")
        
        # Allow processing without selecting specific CP codes since add_total is False by default
        # Users can process the entire file or select specific CP codes as needed

    def sync_collateral_passwords(self, collateral_path):
        """Standalone helper to sync CP passwords from a cash collateral file."""
        collateral_path_clean = (collateral_path or "").strip()
        if not collateral_path_clean:
            raise ValueError("Please select a cash collateral file.")

        new_entries = self._update_passwords_from_collateral(collateral_path_clean)
        return new_entries
    
    def _read_file(self, file_path, **kwargs):
        """Read file with appropriate method based on extension"""
        ext = os.path.splitext(file_path)[1].lower()
        
        try:
            if ext == ".csv":
                df = pd.read_csv(file_path, **kwargs)
            elif ext in [".xls", ".xlsx"]:
                df = pd.read_excel(file_path, **kwargs)
            else:
                raise ValueError(f"Unsupported file type: {ext}")
            
            # Drop rows where all columns are NaN
            df = df.dropna(how='all')
            return df
        except PermissionError:
            self.handle_file_permission_error(file_path, "read")
        except Exception as e:
            if "Permission denied" in str(e) or "being used by another process" in str(e):
                self.handle_file_permission_error(file_path, "read")
            else:
                raise e
    
    def _process_client_position_file(self, file_path, output_path, selected_cp_codes=None, cp_codes_config=None, cash_collateral_path=None):
        """Process client position file and generate report"""
        try:
            # Read the client position file
            df = self._read_file(file_path)
            df.columns = df.columns.str.strip()

            try:
                summary_output_path = os.path.join(output_path, "PS04_Summary.xlsx")
                summary_file = self._create_collateral_summary_excel(file_path, summary_output_path)
                # Log success to a file for debugging
                try:
                    log_file = os.path.join(output_path, "PS04_Summary_log.txt")
                    with open(log_file, "w", encoding="utf-8") as f:
                        f.write(f"PS04 summary Excel generated successfully!\n")
                        f.write(f"Output file: {summary_file}\n")
                except Exception as log_err:
                    pass  # Don't fail if logging fails
            except Exception as summary_error:
                # Log detailed error
                error_msg = f"PS04 summary Excel generation failed!\n"
                error_msg += f"Error: {str(summary_error)}\n"
                error_msg += f"Traceback:\n{traceback.format_exc()}\n"
                self.log_error(output_path, "PS04 summary Excel generation", summary_error)
                # Also write to a separate log file for easier debugging
                try:
                    error_log_file = os.path.join(output_path, "PS04_summary_error.txt")
                    with open(error_log_file, "w", encoding="utf-8") as f:
                        f.write(error_msg)
                except:
                    pass

            new_passwords = 0
            collateral_path_clean = (cash_collateral_path or "").strip()
            if collateral_path_clean:
                try:
                    new_passwords = self._update_passwords_from_collateral(collateral_path_clean)
                except Exception as sync_error:
                    self.log_error(output_path, "Cash collateral password sync", sync_error)

            # Check if BrkrOrCtdnPtcptId column exists
            if "BrkrOrCtdnPtcptId" not in df.columns:
                raise ValueError("Expected a column named 'BrkrOrCtdnPtcptId' in the CSV/Excel file!")
            
            # Filter by selected CP codes if provided
            if selected_cp_codes and len(selected_cp_codes) > 0:
                # Get unique CP codes in the file for diagnostic purposes
                cp_codes_in_file = df['BrkrOrCtdnPtcptId'].astype(str).unique()
                
                df = df[df['BrkrOrCtdnPtcptId'].astype(str).isin(selected_cp_codes)]
                if len(df) == 0:
                    # Return user-friendly message instead of throwing error
                    available_codes = list(cp_codes_in_file)[:10]  # Show first 10 available codes
                    friendly_msg = f"ℹ️ The selected CP code(s) {', '.join(selected_cp_codes)} are not found in the data file.\n\n"
                    friendly_msg += f"📊 Available CP codes in this file: {', '.join(available_codes)}"
                    if len(cp_codes_in_file) > 10:
                        friendly_msg += f" ... (+{len(cp_codes_in_file) - 10} more)"
                    friendly_msg += f"\n\n💡 Total CP codes in file: {len(cp_codes_in_file)}\n"
                    friendly_msg += f"💡 Please select CP codes that exist in your data file."
                    
                    return {
                        "record_count": 0, 
                        "cp_count": 0,
                        "message": friendly_msg,
                        "status": "info"
                    }
            
            # Extract date from filename, or use today's date as fallback
            match = re.search(r'\d{8}', file_path)
            if match:
                extracted_date = match.group()
                today_str = extracted_date
            else:
                # Fallback to today's date if no 8-digit date found in filename
                today_str = datetime.today().strftime("%d%m%Y")
            cp_count = 0
            processed_files = []
            
            # Group by BrkrOrCtdnPtcptId (CP Code column) and process each
            for cp, group in df.groupby("BrkrOrCtdnPtcptId"):
                cp_count += 1
                cp_str = str(cp)
                
                # Get configuration for this CP code using load_passwords function
                try:
                    password_map = load_passwords('master_passwords.json')
                    password = password_map.get(cp_str, '123')
                except Exception:
                    password = '123'
                
                # Load mode and add_total from master_passwords.json FIRST
                mode = '7z'
                add_total = False
                try:
                    with open('master_passwords.json', 'r') as f:
                        all_configs = json.load(f)
                    
                    # Find config for this CP code
                    for item in all_configs:
                        if str(item.get('cp_code', '')) == cp_str:
                            mode = item.get('mode', '7z')
                            add_total = item.get('add_total', False)
                            break
                except Exception:
                    pass
                
                # Override with cp_codes_config if provided (from UI selection)
                if cp_codes_config and cp_str in cp_codes_config:
                    ui_config = cp_codes_config[cp_str]
                    mode = ui_config.get('mode', mode)
                    add_total = ui_config.get('add_total', add_total)
                
                # Calculate specific totals for this CP code
                total_prm_amt = group["PrmAmt"].sum() if "PrmAmt" in group.columns else 0
                total_daly_mtm_val = group["DalyMrkToMktSettlmVal"].sum() if "DalyMrkToMktSettlmVal" in group.columns else 0
                futures_final_val = group["FutrsFnlSttlmVal"].sum() if "FutrsFnlSttlmVal" in group.columns else 0
                exercise_assigned_val = group["ExrcAssgndVal"].sum() if "ExrcAssgndVal" in group.columns else 0
                
                # Combined total
                combined_total = total_prm_amt + total_daly_mtm_val + futures_final_val + exercise_assigned_val
                
                # Create Excel workbook
                wb = Workbook()
                ws = wb.active

                # Write header - keep all original columns as they are
                original_headers = list(group.columns)
                ws.append(original_headers)

                # Find column positions for all needed columns
                col_positions = {}
                for i, col in enumerate(group.columns):
                    col_positions[col] = i
                
                # Convert column index to Excel letter (0->A, 1->B, ..., 37->AL)
                def col_to_excel(col_idx):
                    result = ""
                    while col_idx >= 0:
                        result = chr(col_idx % 26 + 65) + result
                        col_idx = col_idx // 26 - 1
                    return result
                
                # Get Excel column letters for the sum formula
                prm_amt_col = col_to_excel(col_positions.get('PrmAmt', 0)) if 'PrmAmt' in col_positions else None
                daly_mtm_col = col_to_excel(col_positions.get('DalyMrkToMktSettlmVal', 0)) if 'DalyMrkToMktSettlmVal' in col_positions else None
                futrs_col = col_to_excel(col_positions.get('FutrsFnlSttlmVal', 0)) if 'FutrsFnlSttlmVal' in col_positions else None
                exrc_col = col_to_excel(col_positions.get('ExrcAssgndVal', 0)) if 'ExrcAssgndVal' in col_positions else None
                rmks_col = col_to_excel(col_positions.get('Rmks', 0)) if 'Rmks' in col_positions else None

                # Convert Rmks column to object type if we need to add formulas to avoid dtype warnings
                if add_total and 'Rmks' in group.columns and all([prm_amt_col, daly_mtm_col, futrs_col, exrc_col]):
                    group['Rmks'] = group['Rmks'].astype('object')

                # Write rows with formula in existing Rmks column
                row_num = 2  # Excel rows start from 1, header is 1, data starts from 2
                for idx, row in group.iterrows():
                    # If this is the Rmks column, put the formula instead of the value
                    if add_total and 'Rmks' in group.columns and all([prm_amt_col, daly_mtm_col, futrs_col, exrc_col]):
                        formula = f"={prm_amt_col}{row_num}+{daly_mtm_col}{row_num}+{futrs_col}{row_num}+{exrc_col}{row_num}"
                        group.at[idx, 'Rmks'] = formula
                        # IMPORTANT: iterrows() returns a copy; update `row` too so ws.append gets the formula
                        row['Rmks'] = formula
                    
                    ws.append(list(row))
                    row_num += 1

                # If totals required
                if add_total:
                    total_row = [""] * len(group.columns)
                    total_end_row = row_num - 1  # Last data row
                    
                    # Build total formula by summing individual columns directly
                    total_formula_parts = []
                    if prm_amt_col:
                        total_formula_parts.append(f"SUM({prm_amt_col}2:{prm_amt_col}{total_end_row})")
                    if daly_mtm_col:
                        total_formula_parts.append(f"SUM({daly_mtm_col}2:{daly_mtm_col}{total_end_row})")
                    if futrs_col:
                        total_formula_parts.append(f"SUM({futrs_col}2:{futrs_col}{total_end_row})")
                    if exrc_col:
                        total_formula_parts.append(f"SUM({exrc_col}2:{exrc_col}{total_end_row})")
                    
                    # Add TOTAL formula in Rmks column position if it exists, otherwise use ExrcAssgndVal column or last column
                    if total_formula_parts:
                        total_formula = "=" + "+".join(total_formula_parts)
                        
                        # Determine where to place the total
                        if rmks_col:
                            rmks_position = col_positions.get('Rmks', 0)
                            # Add "TOTAL" text in the column just before Rmks
                            if rmks_position > 0:
                                total_row[rmks_position - 1] = "TOTAL"
                            total_row[rmks_position] = total_formula
                        elif exrc_col:
                            # If Rmks doesn't exist, use ExrcAssgndVal column
                            exrc_position = col_positions.get('ExrcAssgndVal', 0)
                            if exrc_position > 0:
                                total_row[exrc_position - 1] = "TOTAL"
                            total_row[exrc_position] = total_formula
                        else:
                            # Fallback: use last column
                            total_row[-2] = "TOTAL"
                            total_row[-1] = total_formula
                    
                    ws.append(total_row)

                # File naming
                base_name = f"FO_PS04_{cp}_{today_str}"
                excel_file = os.path.join(output_path, f"{base_name}.xlsx")

                # Save Excel temp file
                wb.save(excel_file)

                # ZIP mode
                if mode.lower() == "zip":
                    zip_file = os.path.join(output_path, f"{base_name}.zip")
                    with pyzipper.AESZipFile(zip_file, 'w',
                                            compression=pyzipper.ZIP_DEFLATED,
                                            encryption=pyzipper.WZ_AES) as zf:
                        zf.setpassword(password.encode())
                        zf.write(excel_file, os.path.basename(excel_file))
                    processed_files.append(zip_file)

                # 7z mode
                elif mode.lower() in ["gz", "7z"]:
                    gz_file = os.path.join(output_path, f"{base_name}.7z")
                    with py7zr.SevenZipFile(gz_file, 'w', password=password) as archive:
                        archive.write(excel_file, arcname=os.path.basename(excel_file))
                    processed_files.append(gz_file)

                else:
                    raise ValueError(f"Invalid mode '{mode}' for CP Code {cp}. Mode must be 'zip' or '7z'.")

                # Clean up temp Excel file
                os.remove(excel_file)
            
            # Create ZIP with input file and all processed files, then save to database
            self._create_zip_and_save(file_path, processed_files, output_path)
            
            result = {"record_count": len(df), "cp_count": cp_count, "new_passwords": new_passwords}
            
            # Add collateral summary info to result
            if collateral_path_clean:
                collateral_summary_path = os.path.join(output_path, "PS04_Summary.xlsx")
                if os.path.exists(collateral_summary_path):
                    result["collateral_summary"] = collateral_summary_path
            
            return result
            
        except Exception as e:
            self.log_error(output_path, "Error in process_client_position_file", e)
            return None
    
    def _create_zip_and_save(self, input_file, processed_files, output_path):
        """Create ZIP file with input and all processed output files, then save to database"""
        # Create ZIP
        dt = datetime.now().strftime("%Y%m%d_%H%M%S")
        zip_filename = f"CLIENT_POSITION_REPORT_{dt}.zip"
        zip_path = os.path.join(output_path, zip_filename)
        
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            # Add input file
            zipf.write(input_file, os.path.basename(input_file))
            
            # Add all processed output files (7z/zip files for each CP code)
            for processed_file in processed_files:
                if os.path.exists(processed_file):
                    zipf.write(processed_file, os.path.basename(processed_file))
        
        # Insert ZIP into DB (only if db_path is provided)
        if self.db_path:
            with open(zip_path, 'rb') as f:
                zip_blob = f.read()
            
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            insert_report(self.db_path, report_type="CLIENT_POSITION", 
                         created_at=timestamp, modified_at=timestamp, report_blob=zip_blob)
        
        os.remove(zip_path)

    def _update_passwords_from_collateral(self, file_path):
        """Update master_passwords.json with CP/PAN pairs from collateral file."""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Cash collateral file not found: {file_path}")

        df = self._read_file(file_path, header=9)

        if df.empty:
            return 0

        df = df.iloc[:, 1:] if df.shape[1] > 1 else df

        df.columns = [str(col).strip() for col in df.columns]

        required_columns = {"ClientCode", "PANNo"}
        missing = required_columns - set(df.columns)
        if missing:
            raise ValueError(f"Cash collateral file missing required columns: {', '.join(sorted(missing))}")

        df = df.dropna(subset=["ClientCode", "PANNo"], how="any")

        master_path = "master_passwords.json"
        if os.path.exists(master_path):
            with open(master_path, 'r') as fh:
                try:
                    master_data = json.load(fh)
                except json.JSONDecodeError as exc:
                    raise ValueError(f"Invalid JSON in {master_path}: {exc}")
        else:
            master_data = []

        if isinstance(master_data, dict):
            master_data = [
                {
                    'cp_code': self._normalize_identifier(cp_code),
                    'password': self._normalize_identifier((config or {}).get('password')),
                    'mode': (config or {}).get('mode', '7z'),
                    'add_total': (config or {}).get('add_total', False)
                }
                for cp_code, config in master_data.items()
            ]
        elif not isinstance(master_data, list):
            master_data = []

        existing_codes = {
            self._normalize_identifier(entry.get('cp_code')): entry
            for entry in master_data
            if self._normalize_identifier(entry.get('cp_code'))
        }

        new_entries = 0
        for _, row in df.iterrows():
            cp_code = self._normalize_identifier(row.get("ClientCode"))
            password = self._normalize_identifier(row.get("PANNo"))

            if not cp_code or not password:
                continue

            if cp_code in existing_codes:
                continue

            entry = {
                'cp_code': cp_code,
                'password': password,
                'mode': 'zip',
                'add_total': False
            }

            master_data.append(entry)
            existing_codes[cp_code] = entry
            new_entries += 1

        if new_entries:
            with open(master_path, 'w') as fh:
                json.dump(master_data, fh, indent=2)

        return new_entries

    def _create_collateral_summary_excel(self, file_path, output_path):
        """Create aggregated summary Excel from collateral file grouped by BrkrOrCtdnPtcptId."""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Collateral file not found: {file_path}")

        # Read the collateral file (same way as _update_passwords_from_collateral)
        df = self._read_file(file_path)

        if df.empty:
            raise ValueError("Collateral file is empty")

        # Clean column names
        df.columns = [str(col).strip() for col in df.columns]

        # Required columns
        required_columns = ["BrkrOrCtdnPtcptId"]
        sum_columns = [
            "PrmAmt",
            "DalyMrkToMktSettlmVal",
            "FutrsFnlSttlmVal",
            "ExrcAssgndVal"
        ]

        # Check if BrkrOrCtdnPtcptId exists
        if "BrkrOrCtdnPtcptId" not in df.columns:
            raise ValueError(f"Required column 'BrkrOrCtdnPtcptId' not found in file. Available columns: {', '.join(df.columns.tolist())}")

        # Check which sum columns exist
        available_sum_columns = [col for col in sum_columns if col in df.columns]
        if not available_sum_columns:
            raise ValueError(f"None of the required sum columns found. Expected: {', '.join(sum_columns)}. Available: {', '.join(df.columns.tolist())}")

        # Remove rows where BrkrOrCtdnPtcptId is missing
        df = df.dropna(subset=["BrkrOrCtdnPtcptId"])
        df["BrkrOrCtdnPtcptId"] = df["BrkrOrCtdnPtcptId"].astype(str).str.strip()
        df = df[df["BrkrOrCtdnPtcptId"] != ""]

        if df.empty:
            raise ValueError("No valid BrkrOrCtdnPtcptId values found in file")

        # Convert sum columns to numeric, replacing non-numeric values with 0
        for col in available_sum_columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        # Group by BrkrOrCtdnPtcptId and sum the specified columns
        aggregation_dict = {col: 'sum' for col in available_sum_columns}
        grouped_df = df.groupby("BrkrOrCtdnPtcptId", as_index=False).agg(aggregation_dict)

        # Create output directory if it doesn't exist
        output_dir = os.path.dirname(output_path) if os.path.dirname(output_path) else "."
        os.makedirs(output_dir, exist_ok=True)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        # Header including RowTotal
        header = ["BrkrOrCtdnPtcptId"] + available_sum_columns + ["Sum Of Total"]
        ws.append(header)

        # Write aggregated data rows with row totals
        for excel_row_idx, (_, row) in enumerate(grouped_df.iterrows(), start=2):
            row_data = [row["BrkrOrCtdnPtcptId"]]

            # Add numeric values
            for col in available_sum_columns:
                row_data.append(row[col])

            # Create row-total formula
            first_col_idx = 1  # numeric columns start at index 1 (Excel col B)
            last_col_idx = len(available_sum_columns)
            first_letter = self._col_to_excel(first_col_idx)
            last_letter = self._col_to_excel(last_col_idx)

            row_total_formula = f"=SUM({first_letter}{excel_row_idx}:{last_letter}{excel_row_idx})"
            row_data.append(row_total_formula)

            ws.append(row_data)

        # Add TOTAL row (vertical totals)
        total_row = ["TOTAL"]
        last_data_row = len(grouped_df) + 1  # last numeric row number

        # Column totals
        for col in available_sum_columns:
            col_idx = header.index(col)
            col_letter = self._col_to_excel(col_idx)
            total_row.append(f"=SUM({col_letter}2:{col_letter}{last_data_row})")

        # Add total for RowTotal column
        row_total_idx = header.index("Sum Of Total")
        row_total_letter = self._col_to_excel(row_total_idx)
        total_row.append(f"=SUM({row_total_letter}2:{row_total_letter}{last_data_row})")

        ws.append(total_row)

        # Save Excel file
        wb.save(output_path)
        return output_path


    @staticmethod
    def _col_to_excel(col_idx):
        """Convert column index to Excel letter (0->A, 1->B, ..., 25->Z, 26->AA, etc.)"""
        result = ""
        col_num = col_idx + 1  # Convert 0-indexed to 1-indexed (A=1, B=2, ...)
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + 65) + result
            col_num //= 26
        return result

    @staticmethod
    def _normalize_identifier(value):
        if value is None:
            return ""
        text = str(value).strip()
        if not text or text.lower() == 'nan':
            return ""
        return text

    # def protect_existing_excel(self, file_path: str, password: str):
    #     import os
    #     import win32com.client as win32
    #     from win32com.client import constants
    #     import shutil
    #     import tempfile

    #     import os
    #     import win32com.client as win32
    #     from win32com.client import constants

    #     if not os.path.exists(file_path):
    #         raise FileNotFoundError(f"File not found: {file_path}")

    #     temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name

    #     excel = win32.gencache.EnsureDispatch('Excel.Application')
    #     excel.DisplayAlerts = False
    #     excel.ScreenUpdating = False
    #     excel.EnableEvents = False
    #     try:
    #         try:
    #             excel.Calculation = constants.xlCalculationManual
    #         except Exception:
    #             pass

    #         wb = excel.Workbooks.Open(os.path.abspath(file_path), ReadOnly=False)

    #         # SaveAs to temporary file with password
    #         wb.SaveAs(temp_file, FileFormat=51, Password=password)
    #         wb.Close(SaveChanges=False)
    #     finally:
    #         try:
    #             excel.Calculation = constants.xlCalculationAutomatic
    #         except Exception:
    #             pass
    #         excel.DisplayAlerts = True
    #         excel.ScreenUpdating = True
    #         excel.EnableEvents = True
    #         excel.Quit()

    #     # Replace original file with protected version
    #     shutil.move(temp_file, file_path)


class FileComparisonProcessor(BaseProcessor):
    """Processor for File Comparison & Reconciliation (replaced with test.py style comparison)."""

    CPCODE = "CP Code"
    SEGMENT_INDICATOR = "Segment Indicator"
    UCC_CODE = "UCC Code"
    TRADING_MEMBER_PAN = "Trading member PAN"
    ACCOUNT_TYPE = "Account Type"

    def process(self, attachment1_path, attachment2_path, output_path, compare_a_to_b=None, compare_b_to_a=None):
        try:
            self.validate_inputs(
                attachment1_path=attachment1_path,
                attachment2_path=attachment2_path,
                output_path=output_path,
                compare_a_to_b=compare_a_to_b,
                compare_b_to_a=compare_b_to_a
            )
            self.create_output_directory(output_path)

            # Read CSVs without forcing dtype=str to preserve numeric types
            df_system = pd.read_csv(attachment1_path).fillna("")
            df_manual = pd.read_csv(attachment2_path).fillna("")

            system_cols = df_system.columns.tolist()
            manual_cols = df_manual.columns.tolist()

            # Validate headers exactly match
            if system_cols != manual_cols:
                print("❌ Column headers do not match!")
                print("System Columns:", system_cols)
                print("Manual Columns:", manual_cols)
                return

            # Build keyed dicts (handles duplicate keys by suffixing #n)
            system_dict = self._build_keyed_dict(df_system)
            manual_dict = self._build_keyed_dict(df_manual)

            # Ensure output is .xlsx
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"file_comparison_{timestamp}.xlsx"
            output_file = os.path.join(output_path, output_filename)
            if not output_file.lower().endswith(".xlsx"):
                output_file = os.path.splitext(output_file)[0] + ".xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Comparison"
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # Create summary sheet for mismatch row numbers
            ws_summary = wb.create_sheet("Summary")
            ws_summary.append(["KEY", "System Row Number", "Manual Row Number", "Mismatched Columns"])

            # Write a single header row (KEY, SOURCE, then original columns)
            header = ["KEY", "SOURCE"] + system_cols
            ws.append(header)

            mismatches_found = False
            all_keys = sorted(set(system_dict.keys()) | set(manual_dict.keys()))
            summary_data = []

            for key in all_keys:
                row_sys = system_dict.get(key)
                row_man = manual_dict.get(key)

                # Build row value lists (preserve original columns order - manual approach)
                vals_sys = []
                vals_man = []
                mismatched_cols = []
                
                for col in system_cols:
                    # System file - get value as-is without strip()
                    val_sys = row_sys.get(col, "") if row_sys else ""
                    if pd.isna(val_sys):
                        val_sys = ""
                    else:
                        val_sys = str(val_sys)  # Keep original spacing
                    vals_sys.append(val_sys)
                    
                    # Manual file - get value as-is without strip()
                    val_man = row_man.get(col, "") if row_man else ""
                    if pd.isna(val_man):
                        val_man = ""
                    else:
                        val_man = str(val_man)  # Keep original spacing
                    vals_man.append(val_man)

                # Append System row then Manual row (no repeated header)
                file1_excel_row = ws.max_row + 1
                file2_excel_row = file1_excel_row + 1
                
                ws.append([key, "System"] + vals_sys)
                ws.append([key, "Manual"] + vals_man)

                # Highlight differing cells for this key (skip both blank)
                file1_row_idx = file1_excel_row
                file2_row_idx = file2_excel_row
                for i, col in enumerate(system_cols):
                    v1 = vals_sys[i]
                    v2 = vals_man[i]
                    if v1 == "" and v2 == "":
                        continue
                    if v1 != v2:
                        mismatches_found = True
                        mismatched_cols.append(col)
                        data_col_idx = 3 + i  # KEY=1, SOURCE=2, first data col=3
                        ws.cell(row=file1_row_idx, column=data_col_idx).fill = red_fill
                        ws.cell(row=file2_row_idx, column=data_col_idx).fill = red_fill

                # Add to summary if there are mismatches
                if mismatched_cols:
                    summary_data.append([key, file1_excel_row, file2_excel_row, ", ".join(mismatched_cols)])

                # optional spacer row between keys
                ws.append([""] * (2 + len(system_cols)))

            # Write summary data
            for row_data in summary_data:
                ws_summary.append(row_data)

            wb.save(output_file)

            if mismatches_found:
                print(f"✔ Comparison report generated: {output_file}")
                print(f"   - Summary sheet has quick reference to row numbers with mismatches")
            else:
                print("✔ No mismatches found. Files are identical.")

            return {
                'output_file': output_file,
                'summary_count': len(summary_data)
            }

        except Exception as e:
            if output_path:
                self.log_error(output_path, "File Comparison Processing", e)
            raise

    def _get_key(self, row):
        # Primary key
        if row[self.CPCODE] and row[self.SEGMENT_INDICATOR]:
            key = f"{row[self.CPCODE]}_{row[self.SEGMENT_INDICATOR]}"
            print(f"Key generated 1: {key}")
            return key

        # fallback 1
        if row[self.UCC_CODE] and row[self.SEGMENT_INDICATOR]:
            key = f"{row[self.UCC_CODE]}_{row[self.SEGMENT_INDICATOR]}"
            print(f"Key generated 2: {key}")
            return key

        # fallback 2
        if row[self.TRADING_MEMBER_PAN] and row[self.ACCOUNT_TYPE] and row[self.SEGMENT_INDICATOR]:
            key = f"{row[self.TRADING_MEMBER_PAN]}_{row[self.ACCOUNT_TYPE]}_{row[self.SEGMENT_INDICATOR]}"
            print(f"Key generated 3: {key}")
            return key

        return None

    def _build_keyed_dict(self, df):
        keyed = {}
        counts = {}
        for _, row in df.iterrows():
            key = self._get_key(row)
            if not key:
                continue
            # time.sleep(1)  # 1 second sleep
            counts[key] = counts.get(key, 0) + 1
            unique_key = f"{key}#{counts[key]}" if counts[key] > 1 else key
            keyed[unique_key] = row.to_dict()
        return keyed

    def validate_inputs(self, attachment1_path, attachment2_path, output_path, compare_a_to_b, compare_b_to_a):
        """Validate inputs for file comparison."""
        if not attachment1_path or not attachment2_path or not output_path:
            raise ValueError("Please select both attachment files and an output folder before comparing.")
        if not os.path.exists(attachment1_path):
            raise ValueError(f"System file not found:\n{attachment1_path}")
        if not os.path.exists(attachment2_path):
            raise ValueError(f"Manual file not found:\n{attachment2_path}")


class ExerciseAssignmentProcessor(BaseProcessor):
    """Processor for Exercise Assignment Report"""
    
    def process(self, zip_file_path=None, output_path=None, output_format=None, file_paths=None):
        """Process zip/gz/csv/xls/xlsx file(s) for exercise assignment report
        
        Args:
            zip_file_path: Single file path (for backward compatibility)
            output_path: Output directory path
            output_format: List of output formats ['csv', 'xlsx']
            file_paths: List of file paths to process (for multiple files)
        """
        # Handle multiple files
        if file_paths and len(file_paths) > 0:
            return self._process_multiple_files(file_paths, output_path, output_format)
        
        # Handle single file (backward compatibility)
        if not zip_file_path:
            raise ValueError("Please provide either zip_file_path or file_paths")
        
        return self._process_single_file(zip_file_path, output_path, output_format)
    
    def _process_single_file(self, zip_file_path, output_path, output_format=None, skip_info_file=False):
        """Process a single file"""
        try:
            self.validate_inputs(zip_file_path=zip_file_path, output_path=output_path)
            self.create_output_directory(output_path)
            
            # Default to CSV if no format specified
            if output_format is None:
                output_format = ['csv']
            elif not isinstance(output_format, list):
                output_format = [output_format]
            
            file_lower = zip_file_path.lower()
            data_file = None
            extract_dir = None
            
            try:
                # Handle direct CSV/XLS/XLSX files - no extraction needed
                if file_lower.endswith(('.csv', '.xls', '.xlsx')):
                    data_file = zip_file_path
                    # No extraction needed for direct files
                
                # Handle .gz file - extract it first
                elif file_lower.endswith('.gz'):
                    # Create extraction directory for compressed files
                    if extract_dir is None:
                        extract_dir = os.path.join(output_path, "extracted_files")
                        os.makedirs(extract_dir, exist_ok=True)
                    # Extract .gz file
                    extracted_file = self._extract_gz_file(zip_file_path, extract_dir)
                    
                    # Check if extracted file is CSV/XLS/XLSX or another archive
                    extracted_lower = extracted_file.lower()
                    if extracted_lower.endswith(('.csv', '.xls', '.xlsx')):
                        data_file = extracted_file
                    elif extracted_lower.endswith('.zip'):
                        # If .gz contains a .zip, extract that too
                        with zipfile.ZipFile(extracted_file, 'r') as zip_ref:
                            zip_ref.extractall(extract_dir)
                        # Find CSV/XLS/XLSX in the zip contents
                        data_file = self._find_data_file(extract_dir)
                    else:
                        # Try to read the extracted file directly
                        data_file = extracted_file
                
                # Handle .zip file - extract it
                elif file_lower.endswith('.zip'):
                    # Create extraction directory for compressed files
                    if extract_dir is None:
                        extract_dir = os.path.join(output_path, "extracted_files")
                        os.makedirs(extract_dir, exist_ok=True)
                    with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
                        zip_ref.extractall(extract_dir)
                    
                    # Find CSV, XLS, or XLSX file in extracted directory
                    data_file = self._find_data_file(extract_dir)
                
                else:
                    raise ValueError(f"Unsupported file format: {os.path.splitext(zip_file_path)[1]}")
                
            except zipfile.BadZipFile:
                raise ValueError("Invalid zip file. Please ensure the file is a valid zip archive.")
            except Exception as e:
                # Clean up on error
                try:
                    if extract_dir and os.path.exists(extract_dir):
                        shutil.rmtree(extract_dir)
                except:
                    pass
                raise e

            if not data_file or not os.path.exists(data_file):
                if file_lower.endswith(('.csv', '.xls', '.xlsx')):
                    raise ValueError(f"The file {os.path.basename(zip_file_path)} could not be read. Please ensure it is a valid data file.")
                else:
                    raise ValueError("No CSV, XLS, or XLSX file found after extraction. Please ensure the archive contains a valid data file.")
            
            # Read the file into dataframe
            try:
                df = self._read_file(data_file)
                # print(f"DataFrame: {df}")
                # print(f"DataFrame columns: {df.columns}")
                # print(f"DataFrame index: {df.index}")
                # print(f"DataFrame shape: {df.shape}")
                if df is None or df.empty:
                    raise ValueError(f"The file {os.path.basename(data_file)} is empty or could not be read.")
            except Exception as e:
                self.log_error(output_path, f"Error reading {os.path.basename(data_file)}", e)
                raise e
            
            # Filter and segregate data
            # Use original input file name for unique ZIP naming
            original_input_name = os.path.splitext(os.path.basename(zip_file_path))[0]
            result = self._filter_and_segregate(df, data_file, output_path, output_format, original_input_name, skip_info_file=skip_info_file)
            
            # If no records found, return info for consolidated info file
            if result is None:
                # Clean up extracted directory
                if extract_dir:
                    try:
                        if os.path.exists(extract_dir):
                            shutil.rmtree(extract_dir)
                    except:
                        pass
                # Return info structure to indicate no records found (with dataframe for info file)
                return {
                    'file_path': data_file,
                    'filename': os.path.basename(data_file),
                    'dataframe': df,
                    'extract_dir': extract_dir,
                    'output_path': output_path,
                    'zip_file': None,
                    'no_records': True,
                    'input_file_name': original_input_name
                }
            
            # Always create info file for single file processing (if not skipped)
            if not skip_info_file:
                self._create_info_file(output_path, original_input_name, df, has_records=True)
            
            # Clean up extracted directory after successful processing (only if extraction was done)
            if extract_dir:
                try:
                    if os.path.exists(extract_dir):
                        shutil.rmtree(extract_dir)
                except Exception as cleanup_error:
                    # Log cleanup error but don't fail the process
                    self.log_error(output_path, f"Error cleaning up extracted directory: {extract_dir}", cleanup_error)
            
            # Unpack result (zip_path, file_list)
            if result is None:
                zip_file_path = None
                file_list = []
            elif isinstance(result, tuple):
                zip_file_path, file_list = result
            else:
                zip_file_path = result
                file_list = []
            
            return {
                'file_path': data_file,
                'filename': os.path.basename(data_file),
                'dataframe': df,
                'extract_dir': extract_dir,
                'output_path': output_path,
                'zip_file': zip_file_path,
                'zip_path': zip_file_path,  # Add zip_path for email dialog
                'file_list': file_list  # Add file_list for email dialog
            }
                
        except Exception as e:
            # Clean up extracted directory on error
            try:
                if extract_dir and os.path.exists(extract_dir):
                    shutil.rmtree(extract_dir)
            except:
                pass
            self.log_error(output_path, "Exercise Assignment Processing", e)
            raise e
    
    def _read_file_to_df(self, file_path, output_path):
        """Helper to read a file and return dataframe"""
        file_lower = file_path.lower()
        data_file = None
        extract_dir = None
        
        try:
            if file_lower.endswith(('.csv', '.xls', '.xlsx')):
                data_file = file_path
            elif file_lower.endswith('.gz'):
                extract_dir = os.path.join(output_path, "temp_extract")
                os.makedirs(extract_dir, exist_ok=True)
                extracted_file = self._extract_gz_file(file_path, extract_dir)
                if extracted_file.lower().endswith(('.csv', '.xls', '.xlsx')):
                    data_file = extracted_file
                elif extracted_file.lower().endswith('.zip'):
                    with zipfile.ZipFile(extracted_file, 'r') as zip_ref:
                        zip_ref.extractall(extract_dir)
                    data_file = self._find_data_file(extract_dir)
            elif file_lower.endswith('.zip'):
                extract_dir = os.path.join(output_path, "temp_extract")
                os.makedirs(extract_dir, exist_ok=True)
                with zipfile.ZipFile(file_path, 'r') as zip_ref:
                    zip_ref.extractall(extract_dir)
                data_file = self._find_data_file(extract_dir)
            
            if data_file and os.path.exists(data_file):
                df = self._read_file(data_file)
                return df
        finally:
            if extract_dir and os.path.exists(extract_dir):
                try:
                    shutil.rmtree(extract_dir)
                except:
                    pass
        return None
    
    def _process_multiple_files(self, file_paths, output_path, output_format=None):
        """Process multiple files: merge by BrkrOrCtdnPtcptId, Sgmt, Src and create files with date ranges"""
        if not file_paths or len(file_paths) == 0:
            raise ValueError("No files provided for processing")
        
        if not output_path:
            raise ValueError("Please provide output folder before processing")
        
        self.create_output_directory(output_path)
        
        if output_format is None:
            output_format = ['csv']
        elif not isinstance(output_format, list):
            output_format = [output_format]
        
        # Read all files and track which ones were read successfully
        file_dataframes = []  # List of (file_name, dataframe) tuples
        for file_path in file_paths:
            df = self._read_file_to_df(file_path, output_path)
            if df is not None and not df.empty:
                file_name = os.path.splitext(os.path.basename(file_path))[0]
                file_dataframes.append((file_name, df))
        
        if not file_dataframes:
            raise ValueError("No valid data files could be read")
        
        # Combine all dataframes
        all_dfs = [df for _, df in file_dataframes]
        combined_df = pd.concat(all_dfs, ignore_index=True)
        
        # Filter by ExrcAssgndVal != 0
        combined_df['ExrcAssgndVal'] = pd.to_numeric(combined_df['ExrcAssgndVal'], errors='coerce')
        filtered_df = combined_df[(combined_df['ExrcAssgndVal'] != 0) & (combined_df['ExrcAssgndVal'].notna())].copy()
        
        if filtered_df.empty:
            # Create consolidated info file with all file dataframes (no records found)
            self._create_consolidated_info_file(output_path, file_dataframes, has_records=False)
            return {
                'results': [{'file': os.path.basename(fp), 'status': 'info', 'message': 'No records found after merging'} for fp in file_paths],
                'zip_files': [],
                'total_files': len(file_paths),
                'successful': 0,
                'failed': 0
            }
        
        # Update Rmks column
        if 'Rmks' not in filtered_df.columns:
            filtered_df['Rmks'] = ''
        filtered_df['Rmks'] = filtered_df['Rmks'].fillna('').astype(str)
        filtered_df.loc[filtered_df['ExrcAssgndVal'] < 0, 'Rmks'] = 'Assn'
        filtered_df.loc[filtered_df['ExrcAssgndVal'] > 0, 'Rmks'] = 'Ex'
        
        # Remove specified columns
        columns_to_remove = ['PrmAmt', 'DalyMrkToMktSettlmVal', 'FutrsFnlSttlmVal', 'Rsvd1', 'Rsvd2', 'Rsvd3', 'Rsvd4']
        columns_to_drop = [col for col in columns_to_remove if col in filtered_df.columns]
        if columns_to_drop:
            filtered_df = filtered_df.drop(columns=columns_to_drop)
        
        # Normalize columns
        filtered_df['BrkrOrCtdnPtcptId'] = filtered_df['BrkrOrCtdnPtcptId'].fillna("Blank").astype(str).str.strip()
        if 'Sgmt' in filtered_df.columns:
            filtered_df['Sgmt'] = filtered_df['Sgmt'].fillna("").astype(str).str.strip()
        if 'Src' in filtered_df.columns:
            filtered_df['Src'] = filtered_df['Src'].fillna("").astype(str).str.strip()
        
        # Group by BrkrOrCtdnPtcptId, Src, Sgmt
        group_cols = ['BrkrOrCtdnPtcptId']
        if 'Src' in filtered_df.columns:
            group_cols.append('Src')
        if 'Sgmt' in filtered_df.columns:
            group_cols.append('Sgmt')
        
        temp_dir = os.path.join(output_path, "temp_segregated_files")
        os.makedirs(temp_dir, exist_ok=True)
        segregated_files = []
        
        try:
            for group_key, group_df in filtered_df.groupby(group_cols):
                # Get date range from RptgDt
                from_date = to_date = ""
                if 'RptgDt' in group_df.columns:
                    dates = pd.to_datetime(group_df['RptgDt'], errors='coerce').dropna()
                    if not dates.empty:
                        from_date = dates.min().strftime('%Y%m%d')
                        to_date = dates.max().strftime('%Y%m%d')
                
                # Build filename: BrkrOrCtdnPtcptId_Src_Sgmt_from_date_to_date
                brkr_id = str(group_key[0])
                src = str(group_key[1]) if len(group_key) > 1 and 'Src' in group_cols else ""
                sgmt = str(group_key[2]) if len(group_key) > 2 and 'Sgmt' in group_cols else (str(group_key[1]) if len(group_key) > 1 and 'Sgmt' in group_cols else "")
                
                filename_parts = [brkr_id]
                if src:
                    filename_parts.append(src)
                if sgmt:
                    filename_parts.append(sgmt)
                if from_date and to_date:
                    filename_parts.extend([from_date, to_date])
                
                safe_filename = re.sub(r'[<>:"/\\|?*]', '_', '_'.join(filename_parts))
                
                # Create files in selected formats
                for fmt in output_format:
                    if fmt.lower() == 'csv':
                        segregated_file = os.path.join(temp_dir, f"{safe_filename}.csv")
                        group_df.to_csv(segregated_file, index=False)
                        segregated_files.append(segregated_file)
                    elif fmt.lower() == 'xlsx':
                        segregated_file = os.path.join(temp_dir, f"{safe_filename}.xlsx")
                        group_df.to_excel(segregated_file, index=False, engine='openpyxl')
                        segregated_files.append(segregated_file)
            
            # Extract Sgmt value from the data for ZIP filename
            sgmt_value = ""
            if 'Sgmt' in filtered_df.columns:
                # Get unique Sgmt values (non-empty)
                unique_sgmts = filtered_df['Sgmt'].dropna().astype(str).str.strip()
                unique_sgmts = unique_sgmts[unique_sgmts != ""].unique()
                if len(unique_sgmts) > 0:
                    # If all have same Sgmt, use it; otherwise use first one
                    sgmt_value = str(unique_sgmts[0])
                    # Clean up Sgmt value for filename (remove invalid characters)
                    sgmt_value = re.sub(r'[<>:"/\\|?*]', '_', sgmt_value)
            
            # Create zip file with Sgmt in filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if sgmt_value:
                zip_filename = f"ExerciseAssignment_{sgmt_value}_{timestamp}.zip"
            else:
                zip_filename = f"ExerciseAssignment_{timestamp}.zip"
            zip_path = os.path.join(output_path, zip_filename)
            
            # Get list of file names that will be in the ZIP
            file_list = []
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for segregated_file in segregated_files:
                    filename = os.path.basename(segregated_file)
                    zipf.write(segregated_file, filename)
                    file_list.append(filename)
            
            # Create consolidated info file with all file dataframes (records found)
            self._create_consolidated_info_file(output_path, file_dataframes, has_records=True)
            
            # Clean up
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
            
            return {
                'results': [{'file': os.path.basename(fp), 'status': 'success'} for fp in file_paths],
                'zip_files': [zip_path],
                'zip_path': zip_path,  # Add zip_path for email dialog
                'file_list': file_list,  # Add file_list for email dialog
                'total_files': len(file_paths),
                'successful': 1,
                'failed': 0
            }
        except Exception as e:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
            raise e
    
    def _extract_gz_file(self, gz_file_path, extract_dir):
        """Extract .gz file and return path to extracted file"""
        base_name = os.path.basename(gz_file_path)
        # Remove .gz extension
        extracted_name = base_name[:-3] if base_name.lower().endswith('.gz') else base_name
        extracted_path = os.path.join(extract_dir, extracted_name)
        
        with gzip.open(gz_file_path, 'rb') as f_in:
            with open(extracted_path, 'wb') as f_out:
                f_out.write(f_in.read())
        
        return extracted_path
    
    def _find_data_file(self, search_dir):
        """Find CSV, XLS, or XLSX file in directory"""
        for root, dirs, files in os.walk(search_dir):
            for file in files:
                file_ext = os.path.splitext(file)[1].lower()
                if file_ext in ['.csv', '.xls', '.xlsx']:
                    return os.path.join(root, file)
        return None
    
    def _filter_and_segregate(self, df, original_file_path, output_path, output_format=None, input_file_name=None, skip_info_file=False):
        """Filter DataFrame by ExrcAssgndVal > 0, segregate by BrkrOrCtdnPtcptId, and save to zip
        
        Args:
            df: DataFrame to process
            original_file_path: Path to the original data file
            output_path: Output directory path
            output_format: List of output formats ['csv', 'xlsx']
            input_file_name: Base name of the input file (for unique ZIP naming)
            skip_info_file: If True, don't create info file (for multiple files processing)
        """
        # Default to CSV if no format specified
        if output_format is None:
            output_format = ['csv']
        elif not isinstance(output_format, list):
            output_format = [output_format]
        # Validate required columns exist
        required_columns = ['ExrcAssgndVal', 'BrkrOrCtdnPtcptId']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Required columns not found in DataFrame: {missing_columns}. Available columns: {df.columns.tolist()}")
        
        # Filter to exclude rows where ExrcAssgndVal = 0 (keep negative and positive values)
        # Convert ExrcAssgndVal to numeric, handling any non-numeric values
        # Work on a copy to avoid modifying original DataFrame
        filtered_df = df.copy()
        filtered_df['ExrcAssgndVal'] = pd.to_numeric(filtered_df['ExrcAssgndVal'], errors='coerce')
        # Exclude rows where ExrcAssgndVal is 0 or NaN, but keep negative and positive values
        filtered_df = filtered_df[(filtered_df['ExrcAssgndVal'] != 0) & (filtered_df['ExrcAssgndVal'].notna())].copy()
        
        if filtered_df.empty:
            # Create info file only if not skipping (for single file processing)
            if not skip_info_file:
                info_file = self._create_info_file(output_path, input_file_name, df, has_records=False)
            # Return None to indicate no processing needed
            return None
        
        # Update Rmks column based on ExrcAssgndVal value
        # If Rmks column doesn't exist, create it
        if 'Rmks' not in filtered_df.columns:
            filtered_df['Rmks'] = ''
        
        # Convert Rmks column to string type to avoid dtype warnings
        # Replace NaN with empty string first, then convert to string
        filtered_df['Rmks'] = filtered_df['Rmks'].fillna('').astype(str)
        
        # Update Rmks: "Assn" for negative values, "Ex" for positive values
        filtered_df.loc[filtered_df['ExrcAssgndVal'] < 0, 'Rmks'] = 'Assn'
        filtered_df.loc[filtered_df['ExrcAssgndVal'] > 0, 'Rmks'] = 'Ex'
        
        # Remove specified columns from output files
        columns_to_remove = [
            'PrmAmt',
            'DalyMrkToMktSettlmVal',
            'FutrsFnlSttlmVal',
            'Rsvd1',
            'Rsvd2',
            'Rsvd3',
            'Rsvd4'
        ]
        # Only drop columns that exist in the DataFrame
        columns_to_drop = [col for col in columns_to_remove if col in filtered_df.columns]
        if columns_to_drop:
            filtered_df = filtered_df.drop(columns=columns_to_drop)
        
        # Create temporary directory for segregated files
        temp_dir = os.path.join(output_path, "temp_segregated_files")
        os.makedirs(temp_dir, exist_ok=True)
        
        segregated_files = []
        
        try:
            # Segregate by BrkrOrCtdnPtcptId
            # Handle NaN/blank values in BrkrOrCtdnPtcptId
            filtered_df['BrkrOrCtdnPtcptId'] = filtered_df['BrkrOrCtdnPtcptId'].fillna("Blank").astype(str).str.strip()
            filtered_df['BrkrOrCtdnPtcptId'] = filtered_df['BrkrOrCtdnPtcptId'].replace({"": "Blank"})
            
            # Group by BrkrOrCtdnPtcptId
            for brkr_id, group_df in filtered_df.groupby('BrkrOrCtdnPtcptId'):
                # Sanitize filename (remove invalid characters)
                safe_filename = re.sub(r'[<>:"/\\|?*]', '_', str(brkr_id))
                
                # Create files in selected formats
                for fmt in output_format:
                    if fmt.lower() == 'csv':
                        segregated_file = os.path.join(temp_dir, f"{safe_filename}.csv")
                        group_df.to_csv(segregated_file, index=False)
                        segregated_files.append(segregated_file)
                    elif fmt.lower() == 'xlsx':
                        segregated_file = os.path.join(temp_dir, f"{safe_filename}.xlsx")
                        group_df.to_excel(segregated_file, index=False, engine='openpyxl')
                        segregated_files.append(segregated_file)
                    elif fmt.lower() == 'xls':
                        segregated_file = os.path.join(temp_dir, f"{safe_filename}.xls")
                        group_df.to_excel(segregated_file, index=False, engine='xlrd')
                        segregated_files.append(segregated_file)
            
            # Extract Sgmt value from the data for ZIP filename
            sgmt_value = ""
            if 'Sgmt' in filtered_df.columns:
                # Get unique Sgmt values (non-empty)
                unique_sgmts = filtered_df['Sgmt'].dropna().astype(str).str.strip()
                unique_sgmts = unique_sgmts[unique_sgmts != ""].unique()
                if len(unique_sgmts) > 0:
                    # If all have same Sgmt, use it; otherwise use first one
                    sgmt_value = str(unique_sgmts[0])
                    # Clean up Sgmt value for filename (remove invalid characters)
                    sgmt_value = re.sub(r'[<>:"/\\|?*]', '_', sgmt_value)
            
            # Create zip file with Sgmt in filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if sgmt_value:
                zip_filename = f"ExerciseAssignment_{sgmt_value}_{timestamp}.zip"
            elif input_file_name:
                # Sanitize input filename for use in ZIP name
                safe_input_name = re.sub(r'[<>:"/\\|?*]', '_', input_file_name)
                zip_filename = f"ExerciseAssignment_{safe_input_name}_{timestamp}.zip"
            else:
                zip_filename = f"ExerciseAssignment_{timestamp}.zip"
            zip_path = os.path.join(output_path, zip_filename)
            
            # Get list of file names that will be in the ZIP
            file_list = []
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for segregated_file in segregated_files:
                    arcname = os.path.basename(segregated_file)
                    zipf.write(segregated_file, arcname)
                    file_list.append(arcname)
            
            # Clean up temporary directory
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
            
            return zip_path, file_list
            
        except Exception as e:
            # Clean up on error
            try:
                if os.path.exists(temp_dir):
                    shutil.rmtree(temp_dir)
            except:
                pass
            raise e
    
    def _create_info_file(self, output_path, input_file_name, df, has_records=False):
        """Create info file with file name and dates from RptgDt column"""
        info_filename = "exercise_assignment_info.txt"
        info_file_path = os.path.join(output_path, info_filename)
        
        with open(info_file_path, 'w', encoding='utf-8') as f:
            if input_file_name:
                f.write(f"Input File: {input_file_name}\n")
            
            if df is not None and not df.empty and 'RptgDt' in df.columns:
                rptg_dt_values = df['RptgDt'].dropna()
                if len(rptg_dt_values) > 0:
                    unique_dates = sorted(rptg_dt_values.unique())
                    if has_records:
                        f.write(f"Processed dates: {', '.join(str(d) for d in unique_dates)}\n")
                    else:
                        for date_val in unique_dates:
                            f.write(f"On date {date_val} no records found\n")
                else:
                    f.write("No records found\n" if not has_records else "Processed\n")
            else:
                f.write("No records found\n" if not has_records else "Processed\n")
        
        return info_file_path
    
    def _create_consolidated_info_file(self, output_path, file_dataframes, has_records=False):
        """Create consolidated info file for multiple files"""
        info_filename = "exercise_assignment_info.txt"
        info_file_path = os.path.join(output_path, info_filename)
        
        with open(info_file_path, 'w', encoding='utf-8') as f:
            for input_file_name, df in file_dataframes:
                f.write(f"Input File: {input_file_name}\n")
                
                if df is not None and not df.empty and 'RptgDt' in df.columns:
                    rptg_dt_values = df['RptgDt'].dropna()
                    if len(rptg_dt_values) > 0:
                        unique_dates = sorted(rptg_dt_values.unique())
                        if has_records:
                            f.write(f"Processed dates: {', '.join(str(d) for d in unique_dates)}\n")
                        else:
                            for date_val in unique_dates:
                                f.write(f"On date {date_val} no records found\n")
                    else:
                        f.write("No records found\n" if not has_records else "Processed\n")
                else:
                    f.write("No records found\n" if not has_records else "Processed\n")
                
                # Add a blank line between files for readability
                f.write("\n")
        
        return info_file_path
    
    def _read_file(self, file_path):
        """Read file based on extension (CSV, XLS, XLSX, GZ)"""
        file_lower = file_path.lower()
        
        try:
            # Check for .gz files (gzipped files, typically CSV inside)
            if file_lower.endswith('.gz'):
                # Read gzipped file - assume CSV content inside
                df = pd.read_csv(file_path, compression='gzip')
            elif file_lower.endswith('.csv'):
                df = pd.read_csv(file_path)
            elif file_lower.endswith('.xlsx'):
                df = pd.read_excel(file_path, engine='openpyxl')
            elif file_lower.endswith('.xls'):
                df = pd.read_excel(file_path, engine='xlrd')
            else:
                return None
            
            # Drop rows where all columns are NaN
            df = df.dropna(how='all')
            # Strip column names
            df.columns = df.columns.str.strip()
            
            return df
        except PermissionError:
            self.handle_file_permission_error(file_path, "read")
            return None
        except Exception as e:
            if "Permission denied" in str(e) or "being used by another process" in str(e):
                self.handle_file_permission_error(file_path, "read")
                return None
            else:
                raise e
    
    def validate_inputs(self, zip_file_path, output_path):
        """Validate inputs for exercise assignment processing"""
        if not zip_file_path or not output_path:
            raise ValueError("Please select file and output folder before processing.")
        if not os.path.exists(zip_file_path):
            raise ValueError(f"File not found:\n{zip_file_path}")
        file_lower = zip_file_path.lower()
        if not (file_lower.endswith('.zip') or file_lower.endswith('.gz') or 
                file_lower.endswith('.csv') or file_lower.endswith('.xls') or 
                file_lower.endswith('.xlsx')):
            raise ValueError("Selected file must be a .zip, .gz, .csv, .xls, or .xlsx file.")
        
        # Note: 
        # - .zip files should contain CSV, XLS, or XLSX file inside
        # - .gz files are single compressed files (typically CSV content)
        # - .csv, .xls, .xlsx files are processed directly
        # File content validation happens after extraction/processing in the process method